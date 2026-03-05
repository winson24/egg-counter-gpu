import os
# Fix FFmpeg assertion fctx->async_lock failed by disabling multi-threaded decoding
os.environ["OPENCV_FFMPEG_THREADS"] = "1"
# Prioritize DirectShow over Media Foundation for better stability on some Windows systems
os.environ["OPENCV_VIDEOIO_PRIORITY_MSMF"] = "0"

import sys
import argparse
import glob
import time
import threading
import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# ──────────────────────────────────────────────────────────────────────────────
# Path Helpers for PyInstaller / Frozen EXE support
# ──────────────────────────────────────────────────────────────────────────────

def resource_path(relative_path, persistent=False):
    """ 
    Get absolute path to resource, works for dev and for PyInstaller.
    If persistent=True, returns path relative to the EXE directory.
    If persistent=False, returns path relative to the _MEIPASS temp directory.
    """
    if getattr(sys, 'frozen', False):
        if persistent:
            # File stays next to the .exe
            base_path = os.path.dirname(sys.executable)
        else:
            # File is bundled inside the .exe and unzipped to temp folder
            base_path = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
    else:
        base_path = os.path.abspath(".")
        
    return os.path.join(base_path, relative_path)


def initialize_frozen_environment():
    """Register necessary DLL directories for PyTorch/CUDA in frozen environment"""
    if getattr(sys, 'frozen', False):
        try:
            base_path = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
            
            # List of potential DLL search paths (essential for PyTorch/CUDA/NVIDIA)
            potential_paths = [
                os.path.join(base_path, 'torch', 'lib'),
                os.path.join(base_path, 'torch', 'bin'),
                os.path.join(base_path, 'nvidia', 'cuda_runtime', 'bin'),
                os.path.join(base_path, 'nvidia', 'cublas', 'bin'),
                os.path.join(base_path, 'nvidia', 'cudnn', 'bin'),
                os.path.join(base_path, 'nvidia', 'cufft', 'bin'),
                os.path.join(base_path, 'nvidia', 'curand', 'bin'),
                os.path.join(base_path, 'nvidia', 'cusolver', 'bin'),
                os.path.join(base_path, 'nvidia', 'cusparse', 'bin'),
                os.path.join(base_path, 'nvidia', 'nvrtc', 'bin'),
            ]
            
            # Print only existing paths for debugging
            for path in potential_paths:
                if os.path.exists(path):
                    # Add to DLL search path (essential for Python 3.8+)
                    if hasattr(os, 'add_dll_directory'):
                        try:
                            # Note: This tool returns a handle, not a boolean, but try-except is safer
                            os.add_dll_directory(path)
                        except Exception:
                            pass
                    # For compatibility, also add to PATH
                    os.environ['PATH'] = path + os.pathsep + os.environ.get('PATH', '')
            
            print(f"Frozen environment initialized. Added existing DLL directories from: {base_path}")
            
        except Exception as e:
            print(f"Error initializing frozen environment: {e}")

# Run initialization immediately at module load time
initialize_frozen_environment()


import cv2
import numpy as np
from ultralytics import YOLO
import torch

# Log CUDA availability immediately for easier verification
if getattr(sys, 'frozen', False):
    print(f"--- CUDA System Status ---")
    print(f"CUDA Available: {torch.cuda.is_available()}")
    if torch.cuda.is_available():
        print(f"Device Name: {torch.cuda.get_device_name(0)}")
        print(f"Device Count: {torch.cuda.device_count()}")
    else:
        # Check why it might be unavailable
        print(f"PyTorch CUDA support: {torch.backends.cuda.is_built()}")
    print(f"--------------------------")
from pyModbusTCP.client import ModbusClient
from harvesters.core import Harvester
from collections import deque
from PIL import Image, ImageTk
import json
from dataclasses import dataclass
from typing import Optional, Dict, List, Tuple, Any
import queue
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import subprocess
import socketserver
import http.server
import logging

def get_robust_local_ip():
    """Get local IP address using multiple robust methods, prioritizing private network ranges (192, 10, 172)"""
    import socket
    
    # Method 1: Connection test to non-routable address (no internet needed)
    # This usually finds the primary active network interface
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(0.1)
        s.connect(("10.255.255.255", 1))
        ip = s.getsockname()[0]
        s.close()
        if ip and ip != "127.0.0.1" and not ip.startswith("169.254"):
            return ip
    except:
        pass

    # Method 2: Get all IPs from hostname lookup
    try:
        hostname = socket.gethostname()
        _, _, ip_list = socket.gethostbyname_ex(hostname)
        
        # Priority 1: Private IP ranges (192.168.x.x, 10.x.x.x, 172.x.x.x)
        for ip in ip_list:
            if any(ip.startswith(prefix) for prefix in ["192.168.", "10.", "172."]):
                if not ip.startswith("169.254"):
                    return ip
                    
        # Priority 2: Any non-loopback IP
        for ip in ip_list:
            if ip != "127.0.0.1":
                return ip
    except:
        pass

    return "127.0.0.1"

# ──────────────────────────────────────────────────────────────────────────────
# Theme and Constants
# ──────────────────────────────────────────────────────────────────────────────

# Modern Dark Theme Colors
THEME_COLORS = {
    'bg_dark': '#1e1e2e',       # Main window background (Mocha base)
    'bg_card': '#313244',       # Card/Panel background (Surface0)
    'bg_lighter': '#45475a',    # Lighter background (Surface1)
    'text_main': '#cdd6f4',     # Main text (Text)
    'text_dim': '#a6adc8',      # Secondary text (Subtext0)
    'accent': '#89b4fa',        # Primary accent (Blue)
    'accent_hover': '#b4befe',  # Hover state
    'danger': '#f38ba8',        # Red
    'success': '#a6e3a1',       # Green
    'warning': '#f9e2af',       # Yellow
    'border': '#585b70'         # Overlay0
}


def setup_theme(root):
    """Configure modern ttk styles"""
    style = ttk.Style(root)
    # style.theme_use('clam')  # Base theme for better customization

    # Configure base colors
    root.configure(bg=THEME_COLORS['bg_dark'])

    # Standard TFrame
    style.configure('TFrame', background=THEME_COLORS['bg_dark'])

    # Card Frame
    style.configure('Card.TFrame', background=THEME_COLORS['bg_card'],
                    relief='flat', borderwidth=0)

    # Label
    style.configure('TLabel', background=THEME_COLORS['bg_dark'],
                    foreground=THEME_COLORS['text_main'], font=('Segoe UI', 10))
    style.configure('Card.TLabel', background=THEME_COLORS['bg_card'],
                    foreground=THEME_COLORS['text_main'], font=('Segoe UI', 10))
    style.configure('Title.TLabel', background=THEME_COLORS['bg_dark'],
                    foreground=THEME_COLORS['accent'], font=('Segoe UI', 24, 'bold'))
    style.configure('Subtitle.TLabel', background=THEME_COLORS['bg_dark'],
                    foreground=THEME_COLORS['text_dim'], font=('Segoe UI', 11))

    # Entry
    style.configure('TEntry', fieldbackground=THEME_COLORS['bg_lighter'],
                    foreground='black', borderwidth=0, relief='flat', padding=5)

    # Button (Standard)
    style.configure('TButton', font=('Segoe UI', 10, 'bold'),
                    borderwidth=0, focuscolor='none')
    style.map('TButton',
              background=[('active', THEME_COLORS['accent_hover']),
                          ('!disabled', THEME_COLORS['accent'])],
              # Dark text on accent button
              foreground=[('!disabled', '#1e1e2e')])

    # Accent Button (Primary)
    style.configure('Accent.TButton', font=(
        'Segoe UI', 11, 'bold'), borderwidth=0)
    style.map('Accent.TButton',
              background=[('active', THEME_COLORS['accent_hover']),
                          ('!disabled', THEME_COLORS['accent'])],
              foreground=[('!disabled', '#1e1e2e')])

    # Outline/Secondary Button
    style.configure('Secondary.TButton', font=('Segoe UI', 10))
    style.map('Secondary.TButton',
              background=[('active', THEME_COLORS['bg_lighter']),
                          ('!disabled', THEME_COLORS['bg_card'])],
              foreground=[('!disabled', THEME_COLORS['text_main'])])

    # TNotebook
    style.configure(
        'TNotebook', background=THEME_COLORS['bg_dark'], borderwidth=0)
    style.configure('TNotebook.Tab', padding=[12, 8], font=('Segoe UI', 10))
    style.map('TNotebook.Tab',
              background=[('selected', THEME_COLORS['accent']),
                          ('!selected', THEME_COLORS['bg_card'])],
              foreground=[('selected', '#1e1e2e'), ('!selected', THEME_COLORS['text_main'])])

    # TLabelframe
    style.configure('TLabelframe', background=THEME_COLORS['bg_dark'],
                    foreground=THEME_COLORS['text_main'], bordercolor=THEME_COLORS['border'])
    style.configure('TLabelframe.Label', background=THEME_COLORS['bg_dark'],
                    foreground=THEME_COLORS['accent'], font=('Segoe UI', 9, 'bold'))

    return style


# Simple user database
USERS = {
    "admin": {"password": "admin_DAC_123", "role": "admin"},
    "operator": {"password": "operator123", "role": "operator"}
}


@dataclass
class AppSettings:
    """Data class for application settings"""
    model_path: str = ''
    source: str = '0'  # Default to webcam 0
    threshold: float = 0.5
    resolution: str = '640x480'
    device: str = 'auto'
    use_half_precision: bool = False
    plc_ip: str = '169.254.163.50'
    plc_port: int = 502
    plc_address: int = 100
    plc_unit_id: int = 1
    rotate: str = 'none'
    divider_ratio_upper: float = 0.45
    divider_ratio_lower: float = 0.55
    frame_skip: int = 0
    inference_imgsz: int = 640
    show_overlays: bool = True
    persistent_count: int = 0


@dataclass
class Detection:
    """Data class for object detection results"""
    bbox: Tuple[int, int, int, int]
    center: Tuple[int, int]
    classname: str
    confidence: float
    tracker_id: Optional[int] = None
    counted_number: Optional[int] = None


class LoginApp:
    def __init__(self, root):
        self.root = root
        self.root.title("DAC Counting System - Login")
        self.root.geometry("500x550")

        # Setup Theme
        self.style = setup_theme(self.root)

        # Center the window
        self.center_window(500, 550)

        # Main container with gradient-like background (using solid color for Tkinter)
        main_container = ttk.Frame(root)
        main_container.pack(expand=True, fill=tk.BOTH)

        # Card Frame for Login
        login_card = ttk.Frame(
            main_container, style='Card.TFrame', padding="40 40 40 40")
        login_card.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        # Logo/Icon Placeholder (Optional)
        # title_icon = tk.Label(login_card, text="🥚", font=("Segoe UI", 48),
        #                      bg=THEME_COLORS['bg_card'], fg=THEME_COLORS['accent'])
        # title_icon.pack(pady=(0, 10))

        # Title
        tk.Label(login_card, text="DAC SYSTEM",
                 font=("Segoe UI", 24, "bold"),
                 bg=THEME_COLORS['bg_card'], fg=THEME_COLORS['accent']).pack()

        tk.Label(login_card, text="Secure Authentication",
                 font=("Segoe UI", 10),
                 bg=THEME_COLORS['bg_card'], fg=THEME_COLORS['text_dim']).pack(pady=(0, 30))

        # Username Field
        username_frame = tk.Frame(login_card, bg=THEME_COLORS['bg_card'])
        username_frame.pack(fill=tk.X, pady=(0, 15))

        tk.Label(username_frame, text="USERNAME",
                 font=("Segoe UI", 8, "bold"),
                 bg=THEME_COLORS['bg_card'], fg=THEME_COLORS['text_dim']).pack(anchor='w')

        self.username_entry = ttk.Entry(
            username_frame, width=30, font=("Segoe UI", 11))
        self.username_entry.pack(fill=tk.X, ipady=5, pady=(5, 0))

        # Password Field
        password_frame = tk.Frame(login_card, bg=THEME_COLORS['bg_card'])
        password_frame.pack(fill=tk.X, pady=(0, 25))

        tk.Label(password_frame, text="PASSWORD",
                 font=("Segoe UI", 8, "bold"),
                 bg=THEME_COLORS['bg_card'], fg=THEME_COLORS['text_dim']).pack(anchor='w')

        self.password_entry = ttk.Entry(
            password_frame, show="•", width=30, font=("Segoe UI", 11))
        self.password_entry.pack(fill=tk.X, ipady=5, pady=(5, 0))

        # Login Button
        login_btn = ttk.Button(login_card, text="ACCESS SYSTEM", command=self.login,
                               style='Accent.TButton', width=20, cursor="hand2")
        login_btn.pack(fill=tk.X, ipady=5)

        # Footer
        footer_frame = tk.Frame(login_card, bg=THEME_COLORS['bg_card'])
        footer_frame.pack(pady=(20, 0))

        tk.Label(footer_frame, text="Delta Automation Controls",
                 font=("Segoe UI", 8),
                 bg=THEME_COLORS['bg_card'], fg=THEME_COLORS['text_dim']).pack()

        # Bind Enter key to login
        self.root.bind('<Return>', lambda event: self.login())

        # Focus on username field
        self.username_entry.focus()

    def center_window(self, width, height):
        """Center the window on screen"""
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()

        if not username or not password:
            messagebox.showerror(
                "Error", "Please enter both username and password")
            return

        if username in USERS:
            if USERS[username]["password"] == password:
                role = USERS[username]["role"]
                self.open_main_application(role)
            else:
                messagebox.showerror("Error", "Invalid password")
        else:
            messagebox.showerror("Error", "User not found")

    def open_main_application(self, role):
        # Clear login window
        for widget in self.root.winfo_children():
            widget.destroy()

        self.root.title(f"DAC Egg Counting System - {role.capitalize()} Mode")

        # Create main application with role
        self.main_app = EggCountingGUI(self.root, role)


class FrameProcessor:
    """Handles frame processing operations"""

    @staticmethod
    def apply_rotation(frame: np.ndarray, rotation: str) -> np.ndarray:
        """Apply rotation to frame with minimal overhead"""
        if rotation == '90':
            return cv2.rotate(frame, cv2.ROTATE_90_CLOCKWISE)
        elif rotation == '180':
            return cv2.rotate(frame, cv2.ROTATE_180)
        elif rotation == '270':
            return cv2.rotate(frame, cv2.ROTATE_90_COUNTERCLOCKWISE)
        elif rotation == 'flip':
            return cv2.flip(frame, 1)
        return frame

    @staticmethod
    def resize_frame(frame: np.ndarray, resolution: str) -> np.ndarray:
        """Resize frame if needed"""
        if resolution:
            resW, resH = map(int, resolution.split('x'))
            if frame.shape[1] != resW or frame.shape[0] != resH:
                return cv2.resize(frame, (resW, resH))
        return frame

    @staticmethod
    def calculate_center(bbox: Tuple[int, int, int, int]) -> Tuple[int, int]:
        """Calculate center of bounding box efficiently"""
        return ((bbox[0] + bbox[2]) // 2, (bbox[1] + bbox[3]) // 2)

    @staticmethod
    def calculate_distance(point1: Tuple[int, int], point2: Tuple[int, int]) -> float:
        """Calculate Euclidean distance between two points"""
        return np.sqrt((point1[0] - point2[0])**2 + (point1[1] - point2[1])**2)

    @staticmethod
    def calculate_iou(bbox1: Tuple[int, int, int, int], bbox2: Tuple[int, int, int, int]) -> float:
        """Calculate Intersection over Union (IOU) of two bounding boxes."""
        x1, y1, x2, y2 = bbox1
        x3, y3, x4, y4 = bbox2
        
        # Intersection
        xi1 = max(x1, x3)
        yi1 = max(y1, y3)
        xi2 = min(x2, x4)
        yi2 = min(y2, y4)
        
        inter_area = max(0, xi2 - xi1) * max(0, yi2 - yi1)
        
        # Union
        bbox1_area = (x2 - x1) * (y2 - y1)
        bbox2_area = (x4 - x3) * (y4 - y3)
        union_area = bbox1_area + bbox2_area - inter_area
        
        return inter_area / union_area if union_area > 0 else 0


class EggTracker:
    """Handles egg tracking and counting logic"""

    def __init__(self, max_distance: int = 120, persistence_frames: int = 150):
        self.max_distance = max_distance
        self.persistence_frames = persistence_frames
        self.total_eggs_counted = 0
        # Active tracks: {id: (center, bbox, counted, counted_number, classname, last_seen_frame)}
        self.egg_tracking: Dict[int, Tuple[Tuple[int, int], Tuple[int, int, int, int], bool, Optional[int], str, int]] = {}
        # Lost tracks (hidden memory for stability): {id: (center, bbox, counted, counted_number, classname, last_seen_frame)}
        self.lost_tracking: Dict[int, Tuple[Tuple[int, int], Tuple[int, int, int, int], bool, Optional[int], str, int]] = {}
        
        # Spatial Invalidation System: prevents double counting if tracking is lost and regained quickly
        # Format: (center, timestamp, frame_id)
        self.recent_count_locations: List[Tuple[Tuple[int, int], float, int]] = []
        self.count_inhibit_distance = 35  # Higher sensitivity (reduced from 60)
        self.count_inhibit_time = 1.5     # Seconds (reduced from 2.0)
        
        self.frame_counter = 0
        self.next_egg_id = 1
        self.counted_eggs_order = []
        self.class_counts: Dict[str, int] = {}
        self.gui = None  # Reference to GUI for access to FPS

    def update_tracking(self, detections: List[Detection], upper_y: int, lower_y: int) -> int:
        """Update egg tracking and return newly counted eggs."""
        newly_counted = 0
        used_egg_ids = set()
        current_active_tracking = {}
        self.frame_counter += 1

        # 0. Deduplicate overlapping detections (YOLO multiple-detection fix)
        detections = self._deduplicate_detections(detections)
        
        # 1. SORT BY Y (Sequential Count Fix)
        # Ensure eggs are processed from top-to-bottom so counts are always in order
        detections.sort(key=lambda d: d.center[1])

        # 1. Try to match detections to currently active OR recently lost tracks
        for detection in detections:
            center = detection.center
            best_match_id = self._find_best_match(center, used_egg_ids)

            if best_match_id is not None:
                # Retrieve from either active or lost
                if best_match_id in self.egg_tracking:
                    prev_center, prev_bbox, was_counted, old_counted_number, old_classname, _ = self.egg_tracking[best_match_id]
                else:
                    prev_center, prev_bbox, was_counted, old_counted_number, old_classname, _ = self.lost_tracking[best_match_id]
                
                current_counted = was_counted
                current_counted_number = old_counted_number
                current_classname = detection.classname

                # Direct crossing detection (Inside the Zone)
                if not was_counted:
                    curr_y = center[1]
                    
                    # If the egg is between the lines, count it
                    if upper_y <= curr_y <= lower_y:
                        current_counted = True
                        current_counted_number = self.total_eggs_counted + newly_counted + 1
                        current_classname = detection.classname
                        self.class_counts[current_classname] = self.class_counts.get(current_classname, 0) + 1
                        newly_counted += 1
                        self.counted_eggs_order.append(best_match_id)
                        
                        # Store location for Spatial Invalidation
                        self.recent_count_locations.append((center, time.time(), self.frame_counter))

                current_active_tracking[best_match_id] = (
                    center, detection.bbox, current_counted, current_counted_number, current_classname, self.frame_counter)
                used_egg_ids.add(best_match_id)
                
                # Direct link for GUI
                detection.tracker_id = best_match_id
                detection.counted_number = current_counted_number
            else:
                # New egg candidate
                
                # 1. ACTIVE GHOST GUARD:
                # Discard detections that are too close to eggs we are already tracking this frame.
                # This prevents double-boxes from creating duplicate tracks.
                is_active_ghost = False
                for matched_id in used_egg_ids:
                    if matched_id in current_active_tracking:
                        active_data = current_active_tracking[matched_id]
                        active_center = active_data[0]
                        active_bbox = active_data[1]
                        
                        # Use BOTH distance and IOU for ghost detection
                        # A ghost is something VERY close AND overlapping significantly
                        dist = FrameProcessor.calculate_distance(center, active_center)
                        iou = FrameProcessor.calculate_iou(detection.bbox, active_bbox)
                        
                        if dist < 25 or iou > 0.5:
                            is_active_ghost = True
                            break
                if is_active_ghost:
                    continue # Ignore redundant detection
                
                # 2. SPATIAL INVALIDATION CHECK:
                # If a new egg appears exactly where we just counted one, skip it for now.
                current_time = time.time()
                is_duplicate_count = False
                
                # Clean up old locations
                self.recent_count_locations = [loc for loc in self.recent_count_locations 
                                             if current_time - loc[1] < self.count_inhibit_time]
                
                for loc_center, loc_time, loc_frame in self.recent_count_locations:
                    if FrameProcessor.calculate_distance(center, loc_center) < 25: # Reduced threshold
                        is_duplicate_count = True
                        break
                
                track_id = self.next_egg_id
                self.next_egg_id += 1
                
                # IMMEDIATE COUNT LOGIC:
                is_inside = (upper_y <= center[1] <= lower_y)
                is_startup = (self.frame_counter <= 3) # Guard during first 3 frames
                
                current_counted = False
                current_counted_number = None
                
                if is_inside and not is_startup and not is_duplicate_count:
                    # Valid new egg in the zone -> count it immediately
                    current_counted = True
                    current_counted_number = self.total_eggs_counted + newly_counted + 1
                    self.class_counts[detection.classname] = self.class_counts.get(detection.classname, 0) + 1
                    newly_counted += 1
                    self.counted_eggs_order.append(track_id)
                    self.recent_count_locations.append((center, current_time, self.frame_counter))
                elif is_inside and is_startup:
                    # Startup egg: mark as "counted" (processed) without assigning a number
                    # This avoids counting things already on screen when pressing START.
                    current_counted = True
                # Note: if is_duplicate_count is True, we track it but leave current_counted = False
                # This allows it to count later if it moves away from the invalidation spot.
                
                current_active_tracking[track_id] = (
                    center, detection.bbox, current_counted, current_counted_number, detection.classname, self.frame_counter)
                
                # Direct link for GUI
                detection.tracker_id = track_id
                detection.counted_number = current_counted_number
                
                used_egg_ids.add(track_id)

        # 2. Update lost tracking pool
        # Any track that was active but not matched this frame becomes "lost" (hidden memory)
        for egg_id, track_data in self.egg_tracking.items():
            if egg_id not in used_egg_ids:
                self.lost_tracking[egg_id] = track_data

        # Clean up old lost tracks
        self.lost_tracking = {
            eid: data for eid, data in self.lost_tracking.items()
            if self.frame_counter - data[5] < self.persistence_frames
            and eid not in used_egg_ids
        }

        # 3. Finish up
        self.total_eggs_counted += newly_counted
        self.egg_tracking = current_active_tracking
        return newly_counted

    def get_active_tracks(self) -> List[Detection]:
        """Return all active tracks as Detection objects for drawing.
        Lost tracks are held internally for ID persistence but not drawn.
        """
        active_tracks = []
        
        # Add currently active tracks only (this allows visual blinking as requested)
        for egg_id, (center, bbox, counted, count_num, classname, last_seen) in self.egg_tracking.items():
            active_tracks.append(Detection(
                bbox=bbox,
                center=center,
                classname=classname,
                confidence=1.0,
                tracker_id=egg_id,
                counted_number=count_num
            ))
                
        return active_tracks

    def _find_best_match(self, center: Tuple[int, int], used_egg_ids: set) -> Optional[int]:
        """Find the best matching egg ID for a detection in active OR recently lost tracks."""
        min_distance = float('inf')
        best_match_id = None

        # Search active tracking
        for egg_id, (track_center, _, _, _, _, _) in self.egg_tracking.items():
            if egg_id in used_egg_ids:
                continue
            distance = FrameProcessor.calculate_distance(center, track_center)
            if distance < min_distance and distance < self.max_distance:
                min_distance = distance
                best_match_id = egg_id

        # Search lost tracking if no active match found (this prevents blinking/flicker)
        if best_match_id is None:
            for egg_id, (track_center, _, _, _, _, _) in self.lost_tracking.items():
                if egg_id in used_egg_ids:
                    continue
                distance = FrameProcessor.calculate_distance(center, track_center)
                if distance < min_distance and distance < self.max_distance:
                    min_distance = distance
                    best_match_id = egg_id

        return best_match_id

    def _deduplicate_detections(self, detections: List[Detection], min_dist: int = 35) -> List[Detection]:
        """Remove overlapping detections for the same object with high sensitivity."""
        if not detections:
            return []
            
        # Sort by confidence descending
        sorted_dets = sorted(detections, key=lambda x: x.confidence, reverse=True)
        final_dets = []
        
        for det in sorted_dets:
            is_duplicate = False
            for existing in final_dets[:]:
                # Only deduplicate if overlap is HIGH (same egg)
                # If they are just close, keep both (different eggs)
                iou = FrameProcessor.calculate_iou(det.bbox, existing.bbox)
                dist = FrameProcessor.calculate_distance(det.center, existing.center)
                
                if iou > 0.5 or dist < 20: 
                    is_duplicate = True
                    break
            if not is_duplicate:
                final_dets.append(det)
        return final_dets

    def set_initial_count(self, count: int):
        """Set the initial count for persistence"""
        self.total_eggs_counted = max(0, count)

    def reset(self):
        """Reset tracking state"""
        self.total_eggs_counted = 0
        self.class_counts.clear()
        self.egg_tracking.clear()
        self.lost_tracking.clear()
        self.counted_eggs_order.clear()
        self.next_egg_id = 1


class VideoRecorder:
    """Handles video recording functionality with FFmpeg workaround"""

    def __init__(self, segment_duration: int = 300):
        self.segment_duration = segment_duration
        self.recorder = None
        self.start_time = None
        self.segment_counter = 1
        self.is_recording = False
        self.recorder_lock = threading.Lock()  # Add lock for thread safety

    def start_recording(self, frame: np.ndarray) -> bool:
        """Start video recording with safe codec selection"""
        with self.recorder_lock:
            try:
                if self.recorder:
                    self.recorder.release()
                    self.recorder = None

                self.start_time = time.time()
                # Change to MP4
                filename = f"capture_{self._get_datetime_filename()}.mp4"

                # Try mp4v codec for MP4
                codecs = [
                    ('mp4v', cv2.VideoWriter_fourcc(*'mp4v')),
                    ('avc1', cv2.VideoWriter_fourcc(*'avc1')),
                    ('XVID', cv2.VideoWriter_fourcc(*'XVID')), # Fallback
                ]

                successful = False
                for codec_name, fourcc in codecs:
                    try:
                        self.recorder = cv2.VideoWriter(
                            filename,
                            fourcc,
                            20.0,
                            (frame.shape[1], frame.shape[0])
                        )

                        # Test if recorder is actually opened
                        if self.recorder.isOpened():
                            successful = True
                            self.log_message(
                                f"Started recording with {codec_name} codec: {filename}")
                            break
                        else:
                            self.recorder = None
                    except Exception as e:
                        if self.recorder:
                            self.recorder.release()
                            self.recorder = None
                        continue

                if not successful:
                    self.log_message(
                        "Failed to start recording with any codec")
                    return False

                self.is_recording = True
                return True

            except Exception as e:
                self.log_message(f"Error starting recording: {str(e)}")
                if self.recorder:
                    try:
                        self.recorder.release()
                    except:
                        pass
                    self.recorder = None
                return False

    def write_frame(self, frame: np.ndarray) -> bool:
        """Write frame to video and handle segment rotation"""
        with self.recorder_lock:
            if not self.is_recording or not self.recorder:
                return False

            try:
                # Check if segment duration exceeded
                if time.time() - self.start_time >= self.segment_duration:
                    self._rotate_segment(frame)

                # Write frame
                self.recorder.write(frame)
                return True

            except Exception as e:
                self.log_message(f"Error writing frame: {str(e)}")
                # Don't stop recording on single frame error
                return False

    def _rotate_segment(self, frame: np.ndarray):
        """Rotate to new video segment"""
        try:
            if self.recorder:
                self.recorder.release()
                self.recorder = None

            self.segment_counter += 1
            self.start_time = time.time()

            # Start new recording
            filename = f"capture_{self._get_datetime_filename()}_part{self.segment_counter}.mp4"
            fourcc = cv2.VideoWriter_fourcc(*'mp4v')  # Use mp4v codec

            self.recorder = cv2.VideoWriter(
                filename,
                fourcc,
                20.0,
                (frame.shape[1], frame.shape[0])
            )

            if self.recorder.isOpened():
                self.log_message(f"Rotated to new segment: {filename}")
            else:
                self.log_message(f"Failed to create new segment: {filename}")
                self.recorder = None
                self.is_recording = False

        except Exception as e:
            self.log_message(f"Error rotating segment: {str(e)}")
            self.recorder = None
            self.is_recording = False

    def stop_recording(self):
        """Stop video recording safely"""
        with self.recorder_lock:
            self.is_recording = False
            if self.recorder:
                try:
                    self.recorder.release()
                except Exception as e:
                    self.log_message(f"Error releasing recorder: {str(e)}")
                finally:
                    self.recorder = None
            self.log_message("Recording stopped")

    def _get_datetime_filename(self) -> str:
        """Generate filename with current date and time"""
        return datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    def log_message(self, message):
        """Log message if parent GUI exists"""
        if hasattr(self, 'gui') and self.gui:
            self.gui.log_message(message)
        else:
            print(message)


class InferenceEngine:
    """Handles YOLO model inference with optimization"""

    def __init__(self):
        self.model = None
        self.labels = []
        self.device = 'cpu'
        self.half_precision = False
        self.is_loaded = False

    def load_model(self, model_path: str, device: str = 'auto', half_precision: bool = False) -> bool:
        """Load YOLO model with optimization"""
        try:
            if not os.path.exists(model_path):
                # Try to resolve via resource_path if it's a relative path in an EXE
                model_path = resource_path(model_path)
                if not os.path.exists(model_path):
                    self.log_message(f"Model file not found: {model_path}")
                    return False

            if getattr(sys, 'frozen', False):
                self.log_message("Running in a frozen environment (EXE)")
            else:
                self.log_message("Running in a standard Python environment")

            # Determine device
            if device == 'auto':
                self.device = 'cuda' if torch.cuda.is_available() else 'cpu'
            else:
                self.device = device

                self.log_message(f"PyTorch Version: {torch.__version__}")
            self.log_message(f"CUDA Version (Torch): {torch.version.cuda}")
            self.log_message(f"CUDNN Version: {torch.backends.cudnn.version() if torch.backends.cudnn.is_available() else 'N/A'}")
            
            try:
                if torch.cuda.is_available():
                    device_count = torch.cuda.device_count()
                    self.log_message(f"CUDA Available: Yes, Device Count: {device_count}")
                    for i in range(device_count):
                        self.log_message(f"Device {i}: {torch.cuda.get_device_name(i)}")
                else:
                    self.log_message(f"CUDA Available: No")
                    # Try to import ctypes and check for cuda dll presence if needed, but torch.cuda.is_available check is usually sufficient
            except Exception as e:
                self.log_message(f"Error checking CUDA details: {e}")

            # Load model
            self.log_message(f"Loading model from {model_path}...")
            self.model = YOLO(model_path, task='detect')
            self.model.to(self.device)

            self.half_precision = half_precision and self.device.startswith(
                'cuda')

            # Warm up GPU
            if self.device.startswith('cuda'):
                self._warm_up_gpu()

            self.labels = self.model.names
            self.is_loaded = True
            self.log_message(
                f"Model loaded successfully on {self.device.upper()}!")
            return True

        except Exception as e:
            self.log_message(f"Error loading model: {e}")
            return False

    def _warm_up_gpu(self):
        """Warm up GPU with dummy inference"""
        try:
            dummy_input = torch.randn(1, 3, 640, 640).to(self.device)
            if self.half_precision:
                dummy_input = dummy_input.half()
            _ = self.model(dummy_input)
        except Exception as e:
            self.log_message(f"GPU warmup failed: {e}")

    def infer(self, frame: np.ndarray, confidence_threshold: float = 0.5, imgsz: int = 640) -> List[Detection]:
        """Run inference on frame with optimizations"""
        if not self.is_loaded:
            return []

        try:
            # Use optimized inference parameters
            results = self.model(
                frame,
                verbose=False,
                half=self.half_precision,
                imgsz=imgsz,
                device=self.device,
                conf=confidence_threshold,
                iou=0.5  # Add IOU threshold for NMS
            )

            detections = []
            boxes = results[0].boxes

            if boxes is not None:
                for box in boxes:
                    if box.conf.item() > confidence_threshold:
                        xyxy = box.xyxy.cpu().numpy().squeeze().astype(int)
                        center = FrameProcessor.calculate_center(tuple(xyxy))

                        detections.append(Detection(
                            bbox=tuple(xyxy),
                            center=center,
                            classname=self.labels[int(box.cls.item())],
                            confidence=box.conf.item()
                        ))

            return detections

        except Exception as e:
            self.log_message(f"Inference error: {e}")
            return []

    def log_message(self, message):
        """Log message if parent GUI exists"""
        if hasattr(self, 'gui') and self.gui:
            self.gui.log_message(message)
        else:
            print(message)


class CSVExporter:
    def __init__(self, egg_tracker, filename='egg_count_data.csv'):
        self.egg_tracker = egg_tracker
        self.filename = filename
        self.running = False
        self.thread = None
        self.last_count = 0

    def start(self):
        """Start CSV export thread"""
        self.running = True
        self.thread = threading.Thread(target=self._export_loop, daemon=True)
        self.thread.start()

    def stop(self):
        self.running = False
        # Force final write
        try:
            if hasattr(self, 'egg_tracker') and self.egg_tracker:
                current_count = self.egg_tracker.total_eggs_counted
                self._write_csv_row(current_count)
        except Exception as e:
            self.log_message(f"Error writing final CSV row: {e}")

    def _export_loop(self):
        """Export data every minute"""
        while self.running:
            try:
                current_count = self.egg_tracker.total_eggs_counted

                # Only write if count changed or it's time for periodic update
                if current_count != self.last_count or datetime.datetime.now().second == 0:
                    self._write_csv_row(current_count)
                    self.last_count = current_count

            except Exception as e:
                self.log_message(f"CSV export error: {e}")

            time.sleep(60)  # Check every minute

    def _write_csv_row(self, count):
        """Write a single row to CSV"""
        file_exists = os.path.isfile(self.filename)

        with open(self.filename, 'a', newline='') as f:
            writer = csv.writer(f)

            # Write header if file doesn't exist
            if not file_exists:
                writer.writerow([
                    'Timestamp', 'Total_Eggs_Counted', 'Current_Tracked_Objects',
                    'Count_Rate_Per_Minute', 'System_Status', 'FPS'
                ])

            # Calculate count rate
            count_rate = count - self.last_count if self.last_count > 0 else 0

            writer.writerow([
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                count,
                len(self.egg_tracker.egg_tracking),
                count_rate,
                'RUNNING' if hasattr(
                    self.egg_tracker, 'gui') and self.egg_tracker.gui.running else 'STOPPED',
                round(self.egg_tracker.gui.avg_frame_rate, 1) if hasattr(
                    self.egg_tracker, 'gui') else 0
            ])

        self.log_message(
            f"CSV updated: {count} eggs at {datetime.datetime.now().strftime('%H:%M:%S')}")

    def create_daily_report(self):
        """Create a daily summary report"""
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        report_file = f"egg_count_report_{today}.csv"

        try:
            # Read today's data and create summary
            daily_data = []
            if os.path.exists(self.filename):
                with open(self.filename, 'r') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if row['Timestamp'].startswith(today):
                            daily_data.append(row)

            if daily_data:
                total_eggs = int(daily_data[-1]['Total_Eggs_Counted'])
                start_eggs = int(
                    daily_data[0]['Total_Eggs_Counted']) if daily_data else 0
                daily_count = total_eggs - start_eggs

                summary = {
                    'Date': today,
                    'Total_Eggs_Today': daily_count,
                    'Start_Count': start_eggs,
                    'End_Count': total_eggs,
                    # Approximate hours
                    'Production_Hours': len(daily_data) / 60,
                    'Average_Rate_Per_Hour': daily_count / (len(daily_data) / 60) if daily_data else 0
                }

                with open(report_file, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Daily Production Report'])
                    writer.writerow([])
                    for key, value in summary.items():
                        writer.writerow([key, value])

                return report_file

        except Exception as e:
            self.log_message(f"Daily report error: {e}")
        return None

    def log_message(self, message):
        """Log message if parent GUI exists"""
        if hasattr(self.egg_tracker, 'gui') and self.egg_tracker.gui:
            self.egg_tracker.gui.log_message(message)
        else:
            print(message)


class ExcelExporter:
    def __init__(self, egg_tracker, filename='egg_count_live.xlsx'):
        self.egg_tracker = egg_tracker
        self.filename = filename
        self.running = False
        self.thread = None
        self.minute_counter = 0
        self.wb = None
        self.data_ws = None
        self.dashboard_ws = None

    def start(self):
        """Start Excel export thread"""
        self.running = True
        self._initialize_workbook()
        self.thread = threading.Thread(target=self._export_loop, daemon=True)
        self.thread.start()

    def stop(self):
        self.running = False
        self._save_workbook()

    def _initialize_workbook(self):
        """Create or load Excel workbook"""
        try:
            if os.path.exists(self.filename):
                self.wb = openpyxl.load_workbook(self.filename)
            else:
                self.wb = Workbook()
                # Remove default sheet
                self.wb.remove(self.wb.active)

            self._setup_sheets()

        except Exception as e:
            self.log_message(f"Excel initialization error: {e}")
            self.wb = Workbook()
            self._setup_sheets()

    def _setup_sheets(self):
        """Setup data and dashboard sheets"""
        # Data sheet
        if 'Live_Data' not in self.wb.sheetnames:
            self.data_ws = self.wb.create_sheet('Live_Data')
            headers = ['Timestamp', 'Total Eggs', 'Current Tracked',
                       'Count/Minute', 'FPS', 'System Status']
            self.data_ws.append(headers)
            self._format_header(self.data_ws[1])
        else:
            self.data_ws = self.wb['Live_Data']

        # Dashboard sheet
        if 'Dashboard' not in self.wb.sheetnames:
            self.dashboard_ws = self.wb.create_sheet('Dashboard')
            self._create_dashboard()
        else:
            self.dashboard_ws = self.wb['Dashboard']

    def _create_dashboard(self):
        """Create summary dashboard"""
        ws = self.dashboard_ws

        # Title
        ws['A1'] = 'Egg Counting System - Live Dashboard'
        ws['A1'].font = Font(size=16, bold=True)

        # Current stats
        ws['A3'] = 'Current Statistics'
        ws['A3'].font = Font(bold=True)

        dashboard_data = [
            ['Last Update:', '=Live_Data!B2'],
            ['Total Eggs Counted:', '=MAX(Live_Data!B:B)'],
            ['Current Tracked Objects:', '=Live_Data!C2'],
            ['Average Count Rate (per hour):', '=AVERAGE(Live_Data!D:D)*60'],
            ['System FPS:', '=Live_Data!E2'],
            ['System Status:', '=Live_Data!F2']
        ]

        for i, (label, formula) in enumerate(dashboard_data, start=4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = formula
            ws[f'A{i}'].font = Font(bold=True)

    def _format_header(self, row):
        """Format header row"""
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

    def _export_loop(self):
        """Update Excel every minute"""
        while self.running:
            try:
                if datetime.datetime.now().second == 0:  # Update at minute mark
                    self._update_excel()
                    self.minute_counter += 1

                    # Save every 10 minutes to prevent data loss
                    if self.minute_counter % 10 == 0:
                        self._save_workbook()

            except Exception as e:
                self.log_message(f"Excel update error: {e}")

            time.sleep(1)  # Check every second

    def _update_excel(self):
        """Add new row to Excel"""
        if not self.data_ws:
            return

        current_time = datetime.datetime.now()

        # Calculate count rate (eggs per minute)
        if self.data_ws.max_row > 1:
            last_count_cell = self.data_ws.cell(
                row=self.data_ws.max_row, column=2)
            last_count = last_count_cell.value if last_count_cell.value else 0
            count_rate = self.egg_tracker.total_eggs_counted - last_count
        else:
            count_rate = 0

        # Add new data row
        self.data_ws.append([
            current_time.strftime("%Y-%m-%d %H:%M:%S"),
            self.egg_tracker.total_eggs_counted,
            len(self.egg_tracker.egg_tracking),
            count_rate,
            round(self.egg_tracker.gui.avg_frame_rate, 1) if hasattr(
                self.egg_tracker, 'gui') else 0,
            'RUNNING' if hasattr(
                self.egg_tracker, 'gui') and self.egg_tracker.gui.running else 'STOPPED'
        ])

        # Auto-adjust column widths
        for column in self.data_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)  # Cap at 50
            self.data_ws.column_dimensions[column_letter].width = adjusted_width

        self.log_message(
            f"Excel updated at {current_time.strftime('%H:%M:%S')}")

    def _save_workbook(self):
        """Save workbook with error handling"""
        try:
            if self.wb:
                self.wb.save(self.filename)
            return True
        except PermissionError:
            # Try alternative filename if file is open
            temp_filename = f"egg_count_backup_{datetime.datetime.now().strftime('%H%M%S')}.xlsx"
            self.wb.save(temp_filename)
            self.log_message(
                f"Main file busy, saved backup as: {temp_filename}")
            return False
        except Exception as e:
            self.log_message(f"Save error: {e}")
            return False

    def log_message(self, message):
        """Log message if parent GUI exists"""
        if hasattr(self.egg_tracker, 'gui') and self.egg_tracker.gui:
            self.egg_tracker.gui.log_message(message)
        else:
            print(message)


class SimpleMJPEGStreamer:
    """Simple and reliable MJPEG streamer - no FFmpeg required"""

    def __init__(self, port=8080):
        self.port = port
        self.current_frame = None
        self.is_running = False
        self.server = None
        self.thread = None

    def start(self):
        """Start the streamer"""
        if self.is_running:
            return True

        self.is_running = True
        self.thread = threading.Thread(target=self._run_server, daemon=True)
        self.thread.start()

        # Wait for server to start
        time.sleep(2)
        return True

    def stop(self):
        """Stop the streamer"""
        self.is_running = False
        if self.server:
            try:
                self.server.shutdown()
                self.server.server_close()
            except:
                pass

    def update_frame(self, frame):
        """Update frame for streaming"""
        self.current_frame = frame

    def _run_server(self):
        """Run the HTTP server"""
        # Create handler class with streamer reference
        class StreamHandler(http.server.BaseHTTPRequestHandler):
            def do_GET(self):
                if self.path == '/stream.mjpg':
                    self.send_response(200)
                    self.send_header(
                        'Content-type', 'multipart/x-mixed-replace; boundary=frame')
                    self.send_header('Cache-Control', 'no-cache')
                    self.end_headers()

                    try:
                        while self.server.streamer.is_running:
                            if self.server.streamer.current_frame is not None:
                                # Resize for consistent streaming
                                frame = cv2.resize(
                                    self.server.streamer.current_frame, (640, 480))

                                # Encode to JPEG
                                ret, jpeg = cv2.imencode('.jpg', frame, [
                                    cv2.IMWRITE_JPEG_QUALITY, 80
                                ])

                                if ret:
                                    # Standard MJPEG format with correct headers
                                    self.wfile.write(b'--frame\r\n')
                                    self.wfile.write(f"Content-Type: image/jpeg\r\n".encode())
                                    self.wfile.write(f"Content-Length: {len(jpeg)}\r\n\r\n".encode())
                                    self.wfile.write(jpeg.tobytes())
                                    self.wfile.write(b'\r\n')
                                    self.wfile.flush()

                            time.sleep(0.033)  # ~30 FPS

                    except Exception as e:
                        pass  # Client disconnected

                elif self.path == '/' or self.path == '/index.html':
                    self.send_response(200)
                    self.send_header('Content-type', 'text/html')
                    self.end_headers()

                    local_ip = self._get_local_ip()
                    html = f"""
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <title>Delta Monitoring System</title>
                        <meta charset="utf-8">
                        <meta name="viewport" content="width=device-width, initial-scale=1">
                        <style>
                            body {{
                                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                                margin: 0;
                                padding: 0;
                                background: linear-gradient(135deg, #0c2461 0%, #1e3799 100%);
                                color: #333;
                                min-height: 100vh;
                            }}
                            
                            .container {{
                                max-width: 1000px;
                                margin: 0 auto;
                                padding: 20px;
                            }}
                            
                            .header {{
                                background: linear-gradient(90deg, #0066cc 0%, #004c99 100%);
                                color: white;
                                padding: 25px 40px;
                                border-radius: 8px 8px 0 0;
                                text-align: center;
                                box-shadow: 0 4px 12px rgba(0, 0, 0, 0.2);
                            }}
                            
                            .title {{
                                font-size: 28px;
                                font-weight: 600;
                                letter-spacing: 1px;
                                margin-bottom: 5px;
                            }}
                            
                            .subtitle {{
                                font-size: 14px;
                                opacity: 0.9;
                            }}
                            
                            .main-content {{
                                background: #f8f9fa;
                                border-radius: 0 0 8px 8px;
                                padding: 30px;
                                box-shadow: 0 4px 20px rgba(0, 0, 0, 0.1);
                            }}
                            
                            .content-grid {{
                                display: grid;
                                grid-template-columns: 1fr 1fr;
                                gap: 30px;
                                margin-top: 30px;
                            }}
                            
                            @media (max-width: 768px) {{
                                .content-grid {{
                                    grid-template-columns: 1fr;
                                }}
                            }}
                            
                            .panel {{
                                background: white;
                                border-radius: 8px;
                                padding: 25px;
                                box-shadow: 0 2px 8px rgba(0, 0, 0, 0.08);
                            }}
                            
                            .panel-title {{
                                color: #0066cc;
                                font-size: 18px;
                                font-weight: 600;
                                margin-bottom: 20px;
                                padding-bottom: 10px;
                                border-bottom: 2px solid #f0f7ff;
                            }}
                            
                            .info-item {{
                                display: flex;
                                justify-content: space-between;
                                align-items: center;
                                padding: 12px 0;
                                border-bottom: 1px solid #f0f0f0;
                            }}
                            
                            .info-label {{
                                color: #666;
                                font-size: 14px;
                            }}
                            
                            .info-value {{
                                color: #0066cc;
                                font-weight: 600;
                                font-family: 'Courier New', monospace;
                                font-size: 16px;
                            }}
                            
                            .stream-container {{
                                background: #2c3e50;
                                border-radius: 8px;
                                padding: 20px;
                                margin-bottom: 30px;
                                position: relative;
                            }}
                            
                            .stream-header {{
                                display: flex;
                                justify-content: space-between;
                                align-items: center;
                                margin-bottom: 15px;
                                color: white;
                            }}
                            
                            .stream-title {{
                                font-size: 18px;
                                font-weight: 600;
                            }}
                            
                            .stream-frame {{
                                width: 100%;
                                height: 500px;
                                background: #1a252f;
                                border-radius: 4px;
                                overflow: hidden;
                                display: flex;
                                align-items: center;
                                justify-content: center;
                            }}
                            
                            .stream-frame img {{
                                max-width: 100%;
                                max-height: 100%;
                                object-fit: contain;
                            }}
                            
                            .connection-status {{
                                background: #00cc66;
                                color: white;
                                padding: 8px 20px;
                                border-radius: 20px;
                                font-size: 14px;
                                font-weight: 600;
                                display: inline-flex;
                                align-items: center;
                                gap: 8px;
                                margin-top: 20px;
                            }}
                            
                            .connection-status::before {{
                                content: "";
                                width: 10px;
                                height: 10px;
                                background: white;
                                border-radius: 50%;
                                display: inline-block;
                                animation: pulse 2s infinite;
                            }}
                            
                            @keyframes pulse {{
                                0% {{ opacity: 1; }}
                                50% {{ opacity: 0.5; }}
                                100% {{ opacity: 1; }}
                            }}
                            
                            .footer {{
                                text-align: center;
                                margin-top: 30px;
                                padding-top: 20px;
                                border-top: 1px solid #e0e0e0;
                                color: #666;
                                font-size: 12px;
                            }}
                        </style>
                    </head>
                    <body>
                        <div class="container">
                            <div class="header">
                                <div class="title">DELTA MONITORING SYSTEM</div>
                                <div class="subtitle">Real-time Device Monitoring</div>
                            </div>
                            
                            <div class="main-content">
                                <!-- LIVE STREAM AT THE TOP -->
                                <div class="stream-container">
                                    <div class="stream-header">
                                        <div class="stream-title">LIVE STREAM</div>
                                        <div style="color: #aaa; font-size: 14px;">Real-time Video Feed</div>
                                    </div>
                                    <div class="stream-frame">
                                        <img src="/stream.mjpg" alt="Live Stream Feed" id="liveStream">
                                    </div>
                                </div>
                                
                                <!-- DETAILS PANELS AT THE BOTTOM -->
                                <div class="content-grid">
                                    <div class="panel">
                                        <div class="panel-title">DEVICE CONNECTION</div>
                                        <div style="margin-top: 20px;">
                                            <div class="info-item">
                                                <span class="info-label">Local Address:</span>
                                                <span class="info-value">{local_ip}:{self.server.streamer.port}</span>
                                            </div>
                                            <div class="info-item">
                                                <span class="info-label">Network Address:</span>
                                                <span class="info-value">{local_ip}:{self.server.streamer.port}</span>
                                            </div>
                                            <div class="info-item">
                                                <span class="info-label">Connection Status:</span>
                                                <span class="info-value">Connected</span>
                                            </div>
                                            <div class="info-item">
                                                <span class="info-label">Uptime:</span>
                                                <span class="info-value">24:18:05</span>
                                            </div>
                                        </div>
                                        
                                        <div class="connection-status" style="margin-top: 30px;">
                                            Device Online - Streaming Active
                                        </div>
                                    </div>
                                    
                                    <div class="panel">
                                        <div class="panel-title">SYSTEM STATUS</div>
                                        <div style="margin-top: 20px;">
                                            <div class="info-item">
                                                <span class="info-label">Stream Status:</span>
                                                <span class="info-value">Active</span>
                                            </div>
                                            <div class="info-item">
                                                <span class="info-label">Camera Status:</span>
                                                <span class="info-value">Operational</span>
                                            </div>
                                            <div class="info-item">
                                                <span class="info-label">Processing Status:</span>
                                                <span class="info-value">Ready</span>
                                            </div>
                                            <div class="info-item">
                                                <span class="info-label">Last Update:</span>
                                                <span class="info-value" id="lastUpdate">Just now</span>
                                            </div>
                                        </div>
                                    </div>
                                </div>
                                
                                <div class="footer">
                                    Delta Automation Controls © 2025 | Monitoring System v2.0
                                </div>
                            </div>
                        </div>
                        
                        <script>
                            // Update timestamp
                            function updateTimestamp() {{
                                const now = new Date();
                                const timeString = now.toLocaleTimeString();
                                document.getElementById('lastUpdate').textContent = timeString;
                            }}
                            
                            // Update timestamp every minute
                            setInterval(updateTimestamp, 60000);
                            
                            // Initialize timestamp
                            updateTimestamp();
                            
                    """
                    self.wfile.write(html.encode())
                else:
                    self.send_error(404)

            def _get_local_ip(self):
                """Get local IP address using unified robust helper"""
                return get_robust_local_ip()

            def log_message(self, format, *args):
                return  # Suppress logs

        # Create custom server class
        class StreamingServer(socketserver.ThreadingMixIn, http.server.HTTPServer):
            def __init__(self, server_address, RequestHandlerClass, streamer):
                self.streamer = streamer
                super().__init__(server_address, RequestHandlerClass)

        try:
            self.server = StreamingServer(
                ('0.0.0.0', self.port),
                StreamHandler,
                streamer=self
            )

            local_ip = self._get_local_ip()
            self.log_message("🎯 MJPEG Streaming Server Started Successfully!")
            self.log_message(f"📍 Local Access: http://{local_ip}:{self.port}")
            self.log_message(
                f"🌐 Network Access: http://{local_ip}:{self.port}")
            self.log_message("📱 Mobile/Tablet: Use any web browser")

            self.server.serve_forever()

        except Exception as e:
            self.log_message(f"❌ Streaming server error: {e}")
        finally:
            if self.server:
                try:
                    self.server.server_close()
                except:
                    pass
            self.server = None

    def _get_local_ip(self):
        """Get local IP address using unified robust helper"""
        return get_robust_local_ip()

    def log_message(self, message):
        """Log message if parent GUI exists"""
        # This method will be overridden when the streamer is attached to GUI
        print(message)


class EggCountingGUI:
    def __init__(self, root, role):
        self.root = root
        self.role = role
        self.root.title(f"DAC Egg Counting System - {role.capitalize()} Mode")

        # Role-based restrictions
        self.is_admin = (role == "admin")

        # Full screen mode state
        self.full_screen_mode = False  # Start in normal mode

        # Control visibility state
        self.controls_visible = True  # Controls are initially visible

        # Initialize components
        self.settings = AppSettings()
        self.frame_processor = FrameProcessor()
        self.egg_tracker = EggTracker()
        self.video_recorder = VideoRecorder()
        self.inference_engine = InferenceEngine()

        # Add exporters and streamers
        self.csv_exporter = CSVExporter(self.egg_tracker)
        self.excel_exporter = ExcelExporter(self.egg_tracker)
        self.mjpeg_streamer = SimpleMJPEGStreamer(
            port=8080)  # Use the simple version

        # Link components to GUI for logging
        self.egg_tracker.gui = self
        self.video_recorder.gui = self
        self.inference_engine.gui = self
        self.csv_exporter.egg_tracker.gui = self
        self.excel_exporter.egg_tracker.gui = self

        # Override streamer log_message to use GUI
        self.mjpeg_streamer.log_message = self.log_message

        # State variables
        self.running = False
        self.model_loaded = False
        self.loading_model = False # Track model loading state
        self.adjustment_mode = False
        self.plc_connection_attempted = False

        # Pause state for spacebar functionality
        self.counting_paused = False
        self.pause_start_time = None
        self.total_pause_time = 0
        self.pause_lock = threading.Lock()  # Add lock for thread safety

        # Camera sources
        self.cap = None
        self.delta_cam = None
        self.current_source_type = None  # 'camera', 'image', 'video', 'folder'

        # PLC
        self.plc = None
        self.last_plc_count = 0
        self.last_plc_update_time = 0
        self.plc_update_interval = 2

        # Performance tracking
        self.frame_counter = 0
        self.avg_frame_rate = 0.0
        self.frame_rate_buffer = deque(maxlen=100)

        # GUI state
        self.current_frame = None
        self.photo = None

        # Threading
        self.video_thread = None
        # Limit queue size to prevent memory buildup
        self.frame_queue = queue.Queue(maxsize=1)

        # File sources
        self.image_files = []
        self.current_image_index = 0
        self.video_file = None

        self.load_settings()
        self.create_widgets()
        self._update_button_states() # Initialize button states
        self.setup_bindings()

        # Handle window close safely
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        # Display welcome message
        self.log_message(f"Welcome {self.role.capitalize()}!")
        self.log_message(
            f"Role-based access: {'Full administrative access' if self.is_admin else 'Operator access with full control tab'}")
        self.log_message(
            "Note: Operators have full access to all features in the Control tab")

    def load_settings(self):
        """Load settings from config file"""
        config_path = resource_path('egg_counter_config.json', persistent=True)
            
        try:
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    saved_settings = json.load(f)
                    for key, value in saved_settings.items():
                        if hasattr(self.settings, key):
                            setattr(self.settings, key, value)
                    
                    # Apply persistent count to tracker immediately if model not running
                    if not self.running:
                        self.egg_tracker.set_initial_count(self.settings.persistent_count)
        except Exception as e:
            self.log_message(f"Error loading settings: {e}")

    def save_settings(self):
        """Save settings to config file"""
        config_path = resource_path('egg_counter_config.json', persistent=True)
            
        try:
            # Update persistent count before saving
            self.settings.persistent_count = self.egg_tracker.total_eggs_counted
            
            settings_dict = {k: v for k, v in self.settings.__dict__.items()
                             if not k.startswith('_')}
            with open(config_path, 'w') as f:
                json.dump(settings_dict, f, indent=4)
        except Exception as e:
            self.log_message(f"Error saving settings: {e}")

    def create_widgets(self):
        """Create optimized GUI layout"""
        # Main container with PanedWindow for resizable sections
        self.main_container = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        self.main_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Left panel - Controls (resizable)
        self.left_frame = ttk.Frame(
            self.main_container, width=380, style='Card.TFrame', padding=10)
        self.main_container.add(self.left_frame, weight=1)

        # Right panel - Video and logs (resizable)
        # apply style for consistency
        self.right_frame = ttk.Frame(self.main_container, style='Card.TFrame')
        self.main_container.add(self.right_frame, weight=3)

        # Build components
        self.build_control_panel(self.left_frame)
        self.build_video_display(self.right_frame)
        self.build_status_bar()

    def build_status_bar(self):
        """Build status bar"""
        self.status_frame = ttk.Frame(self.root, style='Card.TFrame')
        self.status_frame.pack(fill=tk.X, side=tk.BOTTOM,
                               padx=10, pady=(0, 10))

        # Helper to create status items
        def create_status_item(parent, var, width=20):
            frame = ttk.Frame(parent, style='Card.TFrame')
            frame.pack(side=tk.LEFT, padx=5, pady=5)
            # Just use label
            ttk.Label(frame, textvariable=var, style='Card.TLabel',
                      font=('Segoe UI', 9)).pack()

        # System status
        self.system_status_var = tk.StringVar(
            value=f"System Stopped - {self.role.capitalize()} Mode")
        create_status_item(self.status_frame, self.system_status_var, 30)

        # Device status
        self.device_status_var = tk.StringVar(value="Device: Not Loaded")
        create_status_item(self.status_frame, self.device_status_var)

        # Camera status
        self.camera_status_var = tk.StringVar(value="Camera: Disconnected")
        create_status_item(self.status_frame, self.camera_status_var)

        # Pause status in status bar
        self.pause_status_var = tk.StringVar(value="Counting: Active")
        create_status_item(self.status_frame, self.pause_status_var)

        # Logout button in status bar for all users
        self.logout_btn_status = ttk.Button(self.status_frame, text="Logout",
                                            command=self.logout, width=8, style='Secondary.TButton')
        self.logout_btn_status.pack(side=tk.RIGHT, padx=10, pady=5)

    def build_control_panel(self, parent):
        """Build optimized control panel based on user role"""
        self.notebook = ttk.Notebook(parent)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # Create tabs based on user role
        if self.is_admin:
            # Admin sees all tabs
            model_tab = ttk.Frame(self.notebook)
            plc_tab = ttk.Frame(self.notebook)
            advanced_tab = ttk.Frame(self.notebook)
            controls_tab = ttk.Frame(self.notebook)

            self.notebook.add(model_tab, text="Model & Source")
            self.notebook.add(plc_tab, text="PLC Settings")
            self.notebook.add(advanced_tab, text="Advanced")
            self.notebook.add(controls_tab, text="Controls")

            # Build all tabs
            self.build_model_tab(model_tab)
            self.build_plc_tab(plc_tab)
            self.build_advanced_tab(advanced_tab)
            self.build_controls_tab(controls_tab)

            # Default to Controls tab
            # Select last tab (Controls)
            self.notebook.select(len(self.notebook.tabs())-1)

        else:
            # Operator sees only Controls tab with all features
            controls_tab = ttk.Frame(self.notebook)
            self.notebook.add(controls_tab, text="Controls")
            self.build_operator_controls_tab(controls_tab)

    def build_model_tab(self, parent):
        """Build model and source configuration tab"""
        # Model selection
        ttk.Label(parent, text="YOLO Model Path:", font=("Segoe UI", 9, "bold")).grid(
            row=0, column=0, sticky=tk.W, pady=5, padx=5)
        model_frame = ttk.Frame(parent)
        model_frame.grid(row=0, column=1, sticky=tk.EW, pady=5)
        self.model_path_var = tk.StringVar(value=self.settings.model_path)
        self.model_entry = ttk.Entry(
            model_frame, textvariable=self.model_path_var, width=25)
        self.model_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        browse_btn = ttk.Button(
            model_frame, text="Browse", command=self.browse_model, width=8)
        browse_btn.pack(side=tk.RIGHT, padx=(5, 0))

        # Source selection
        ttk.Label(parent, text="Video Source:", font=("Segoe UI", 9, "bold")).grid(
            row=1, column=0, sticky=tk.W, pady=5)
        source_frame = ttk.Frame(parent)
        source_frame.grid(row=1, column=1, sticky=tk.EW, pady=5)
        self.source_var = tk.StringVar(value=self.settings.source)
        source_combo = ttk.Combobox(
            source_frame, textvariable=self.source_var, width=20)
        source_combo['values'] = (
            '0', '1', '2', 'deltacam', 'Image File', 'Image Folder', 'Video File')
        source_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        source_browse_btn = ttk.Button(
            source_frame, text="Browse", command=self.browse_source, width=8)
        source_browse_btn.pack(side=tk.RIGHT, padx=(5, 0))

        # Confidence threshold
        ttk.Label(parent, text="Confidence Threshold:", font=("Segoe UI", 9, "bold")).grid(
            row=2, column=0, sticky=tk.W, pady=5)
        threshold_frame = ttk.Frame(parent)
        threshold_frame.grid(row=2, column=1, sticky=tk.EW, pady=5)
        self.threshold_var = tk.DoubleVar(value=self.settings.threshold)
        threshold_scale = ttk.Scale(threshold_frame, from_=0.1, to=1.0, variable=self.threshold_var,
                                    orient=tk.HORIZONTAL, length=150)
        threshold_scale.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.threshold_label = ttk.Label(
            threshold_frame, text=f"{self.threshold_var.get():.2f}", width=5)
        self.threshold_label.pack(side=tk.RIGHT, padx=(5, 0))

        # Update threshold label when scale changes
        def update_threshold_label(*args):
            self.threshold_label.config(text=f"{self.threshold_var.get():.2f}")
        self.threshold_var.trace('w', update_threshold_label)

        # Resolution
        ttk.Label(parent, text="Resolution:", font=("Segoe UI", 9, "bold")).grid(
            row=3, column=0, sticky=tk.W, pady=5)
        self.resolution_var = tk.StringVar(value=self.settings.resolution)
        res_combo = ttk.Combobox(
            parent, textvariable=self.resolution_var, width=30)
        res_combo['values'] = ('640x480', '800x600',
                               '1024x768', '1280x720', '1920x1080')
        res_combo.grid(row=3, column=1, sticky=tk.EW, pady=5)

        # Device selection
        ttk.Label(parent, text="Inference Device:", font=("Segoe UI", 9, "bold")).grid(
            row=4, column=0, sticky=tk.W, pady=5)
        self.device_var = tk.StringVar(value=self.settings.device)
        device_combo = ttk.Combobox(
            parent, textvariable=self.device_var, width=30)
        device_combo['values'] = ('auto', 'cpu', 'cuda', 'cuda:0')
        device_combo.grid(row=4, column=1, sticky=tk.EW, pady=5)

        # Inference Image Size (imgsz)
        ttk.Label(parent, text="Inference Size (imgsz):", font=("Segoe UI", 9, "bold")).grid(
            row=5, column=0, sticky=tk.W, pady=5)
        self.imgsz_var = tk.IntVar(value=self.settings.inference_imgsz)
        imgsz_combo = ttk.Combobox(parent, textvariable=self.imgsz_var, width=30)
        imgsz_combo['values'] = (320, 416, 512, 640)
        imgsz_combo.grid(row=5, column=1, sticky=tk.EW, pady=5)

        # Half precision and Overlays
        cb_frame = ttk.Frame(parent)
        cb_frame.grid(row=6, column=0, columnspan=2, sticky=tk.W, pady=5)

        self.half_precision_var = tk.BooleanVar(
            value=self.settings.use_half_precision)
        half_precision_cb = ttk.Checkbutton(
            cb_frame, text="Use Half Precision (FP16)", variable=self.half_precision_var)
        half_precision_cb.pack(side=tk.LEFT, padx=(0, 20))

        self.show_overlays_var = tk.BooleanVar(value=self.settings.show_overlays)
        show_overlays_cb = ttk.Checkbutton(
            cb_frame, text="Show Detections/Overlays", variable=self.show_overlays_var)
        show_overlays_cb.pack(side=tk.LEFT)

    def build_plc_tab(self, parent):
        """Build PLC configuration tab"""
        # PLC IP
        ttk.Label(parent, text="PLC IP Address:").grid(
            row=0, column=0, sticky=tk.W, pady=5)
        self.plc_ip_var = tk.StringVar(value=self.settings.plc_ip)
        plc_ip_entry = ttk.Entry(
            parent, textvariable=self.plc_ip_var, width=30)
        plc_ip_entry.grid(row=0, column=1, sticky=tk.EW, pady=5)

        # PLC Port
        ttk.Label(parent, text="PLC Port:").grid(
            row=1, column=0, sticky=tk.W, pady=5)
        self.plc_port_var = tk.IntVar(value=self.settings.plc_port)
        plc_port_entry = ttk.Entry(
            parent, textvariable=self.plc_port_var, width=30)
        plc_port_entry.grid(row=1, column=1, sticky=tk.EW, pady=5)

        # PLC Address
        ttk.Label(parent, text="Register Address:").grid(
            row=2, column=0, sticky=tk.W, pady=5)
        self.plc_address_var = tk.IntVar(value=self.settings.plc_address)
        plc_address_entry = ttk.Entry(
            parent, textvariable=self.plc_address_var, width=30)
        plc_address_entry.grid(row=2, column=1, sticky=tk.EW, pady=5)

        # Unit ID
        ttk.Label(parent, text="Unit ID:").grid(
            row=3, column=0, sticky=tk.W, pady=5)
        self.plc_unit_id_var = tk.IntVar(value=self.settings.plc_unit_id)
        plc_unit_id_entry = ttk.Entry(
            parent, textvariable=self.plc_unit_id_var, width=30)
        plc_unit_id_entry.grid(row=3, column=1, sticky=tk.EW, pady=5)

        # PLC Status
        ttk.Label(parent, text="PLC Status:").grid(
            row=4, column=0, sticky=tk.W, pady=5)
        self.plc_status_var = tk.StringVar(value="Disconnected")
        ttk.Label(parent, textvariable=self.plc_status_var, foreground="red").grid(
            row=4, column=1, sticky=tk.W, pady=5)

        # PLC Controls
        plc_controls = ttk.Frame(parent)
        plc_controls.grid(row=5, column=0, columnspan=2, pady=10)
        connect_btn = ttk.Button(
            plc_controls, text="Connect PLC", command=self.connect_plc)
        connect_btn.pack(side=tk.LEFT, padx=(0, 5))
        disconnect_btn = ttk.Button(
            plc_controls, text="Disconnect PLC", command=self.disconnect_plc)
        disconnect_btn.pack(side=tk.LEFT, padx=(0, 5))
        test_btn = ttk.Button(
            plc_controls, text="Test Communication", command=self.test_plc)
        test_btn.pack(side=tk.LEFT)

        parent.columnconfigure(1, weight=1)

    def build_advanced_tab(self, parent):
        """Build advanced settings tab"""
        # Rotation
        ttk.Label(parent, text="Rotation:").grid(
            row=0, column=0, sticky=tk.W, pady=5)
        self.rotate_var = tk.StringVar(value=self.settings.rotate)
        rotate_combo = ttk.Combobox(
            parent, textvariable=self.rotate_var, width=30)
        rotate_combo['values'] = ('none', '90', '180', '270', 'flip')
        rotate_combo.grid(row=0, column=1, sticky=tk.EW, pady=5)

        # Upper divider line
        ttk.Label(parent, text="Upper Line Ratio:").grid(
            row=1, column=0, sticky=tk.W, pady=5)
        upper_frame = ttk.Frame(parent)
        upper_frame.grid(row=1, column=1, sticky=tk.EW, pady=5)
        self.divider_upper_var = tk.DoubleVar(
            value=self.settings.divider_ratio_upper)
        upper_scale = ttk.Scale(upper_frame, from_=0.05, to=0.94,
                                variable=self.divider_upper_var,
                                orient=tk.HORIZONTAL, length=150)
        upper_scale.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.upper_label = ttk.Label(
            upper_frame, text=f"{self.divider_upper_var.get():.2f}", width=5)
        self.upper_label.pack(side=tk.RIGHT, padx=(5, 0))

        def update_upper(*args):
            val = self.divider_upper_var.get()
            # Keep upper below lower
            if val >= self.settings.divider_ratio_lower:
                val = self.settings.divider_ratio_lower - 0.01
                self.divider_upper_var.set(val)
            self.settings.divider_ratio_upper = val
            self.upper_label.config(text=f"{val:.2f}")
        self.divider_upper_var.trace('w', update_upper)

        # Lower divider line
        ttk.Label(parent, text="Lower Line Ratio:").grid(
            row=2, column=0, sticky=tk.W, pady=5)
        lower_frame = ttk.Frame(parent)
        lower_frame.grid(row=2, column=1, sticky=tk.EW, pady=5)
        self.divider_lower_var = tk.DoubleVar(
            value=self.settings.divider_ratio_lower)
        lower_scale = ttk.Scale(lower_frame, from_=0.06, to=0.95,
                                variable=self.divider_lower_var,
                                orient=tk.HORIZONTAL, length=150)
        lower_scale.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.lower_label = ttk.Label(
            lower_frame, text=f"{self.divider_lower_var.get():.2f}", width=5)
        self.lower_label.pack(side=tk.RIGHT, padx=(5, 0))

        def update_lower(*args):
            val = self.divider_lower_var.get()
            # Keep lower above upper
            if val <= self.settings.divider_ratio_upper:
                val = self.settings.divider_ratio_upper + 0.01
                self.divider_lower_var.set(val)
            self.settings.divider_ratio_lower = val
            self.lower_label.config(text=f"{val:.2f}")
        self.divider_lower_var.trace('w', update_lower)

        # Frame skip
        ttk.Label(parent, text="Frame Skip:").grid(
            row=3, column=0, sticky=tk.W, pady=5)
        self.frame_skip_var = tk.IntVar(value=self.settings.frame_skip)
        frame_skip_combo = ttk.Combobox(
            parent, textvariable=self.frame_skip_var, width=30)
        frame_skip_combo['values'] = (0, 1, 2, 3, 4, 5)
        frame_skip_combo.grid(row=3, column=1, sticky=tk.EW, pady=5)

        # Max distance for tracking
        ttk.Label(parent, text="Max Tracking Distance:").grid(
            row=4, column=0, sticky=tk.W, pady=5)
        self.max_distance_var = tk.IntVar(value=self.egg_tracker.max_distance)
        max_distance_entry = ttk.Entry(
            parent, textvariable=self.max_distance_var, width=30)
        max_distance_entry.grid(row=4, column=1, sticky=tk.EW, pady=5)

        # PLC update interval
        ttk.Label(parent, text="PLC Update Interval (s):").grid(
            row=5, column=0, sticky=tk.W, pady=5)
        self.plc_update_interval_var = tk.IntVar(
            value=self.plc_update_interval)
        plc_update_entry = ttk.Entry(parent, textvariable=self.plc_update_interval_var,
                                     width=30)
        plc_update_entry.grid(row=5, column=1, sticky=tk.EW, pady=5)

        parent.columnconfigure(1, weight=1)

    def build_controls_tab(self, parent):
        """Build controls tab for admin"""

        # Logout button frame (Fixed at bottom)
        logout_frame = ttk.Frame(parent)
        logout_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10, padx=10)

        logout_btn = ttk.Button(logout_frame, text="⏻ Logout",
                                command=self.logout, style='Danger.TButton', width=20)
        logout_btn.pack(side=tk.BOTTOM)

        # Scrollable container for controls
        canvas = tk.Canvas(
            parent, bg=THEME_COLORS['bg_dark'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(
            parent, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Helper for sections
        def create_section(title):
            frame = ttk.LabelFrame(scrollable_frame, text=title, padding=10)
            frame.pack(fill=tk.X, padx=10, pady=5)
            return frame

        # Main controls
        controls_frame = create_section("System Control")

        self.start_btn = ttk.Button(controls_frame, text="▶ Start System",
                                    command=self.start_system, width=15, style='Accent.TButton')
        self.start_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.stop_btn = ttk.Button(controls_frame, text="⏹ Stop System",
                                   command=self.stop_system, width=15)
        self.stop_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.load_model_btn = ttk.Button(controls_frame, text="📂 Load Model",
                                         command=self.load_model, width=15)
        self.load_model_btn.pack(side=tk.LEFT)

        # Counting controls
        counting_frame = create_section("Counting Operations")

        self.reset_btn = ttk.Button(counting_frame, text="↺ Reset Count",
                                    command=self.reset_count, width=15)
        self.reset_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.adjust_btn = ttk.Button(counting_frame, text="📏 Adjust Divider",
                                     command=self.toggle_adjustment, width=15)
        self.adjust_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.capture_btn = ttk.Button(counting_frame, text="📷 Capture Image",
                                      command=self.capture_image, width=15)
        self.capture_btn.pack(side=tk.LEFT)

        # Performance controls
        perf_frame = create_section("Performance & Display")

        self.toggle_device_btn = ttk.Button(perf_frame, text="⚡ Toggle GPU/CPU",
                                            command=self.toggle_device, width=15)
        self.toggle_device_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.save_settings_btn = ttk.Button(perf_frame, text="💾 Save Settings",
                                            command=self.save_settings, width=15)
        self.save_settings_btn.pack(side=tk.LEFT)

        # Full screen toggle button
        ttk.Button(perf_frame, text="⛶ Toggle Full Screen",
                   command=self.toggle_full_screen, width=20).pack(side=tk.LEFT, padx=(10, 5))

        # Excel Export Controls
        excel_frame = create_section("Data Export")

        self.open_excel_btn = ttk.Button(excel_frame, text="📊 Open Excel",
                                         command=self.open_excel_file, width=12)
        self.open_excel_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.daily_report_btn = ttk.Button(excel_frame, text="📅 Daily Report",
                                           command=self.create_daily_report, width=12)
        self.daily_report_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.export_status_btn = ttk.Button(excel_frame, text="ℹ Status",
                                            command=self.show_export_status, width=12)
        self.export_status_btn.pack(side=tk.LEFT)

        # Streaming Controls
        stream_frame = create_section("Wireless Streaming")

        self.start_stream_btn = ttk.Button(stream_frame, text="📡 Start Stream",
                                           command=self.start_streaming, width=15)
        self.start_stream_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.stop_stream_btn = ttk.Button(stream_frame, text="⏹ Stop Stream",
                                          command=self.stop_streaming, width=15)
        self.stop_stream_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.stream_info_btn = ttk.Button(stream_frame, text="ℹ Info",
                                          command=self.show_stream_info, width=12)
        self.stream_info_btn.pack(side=tk.LEFT)

        # Statistics display
        stats_frame = ttk.LabelFrame(scrollable_frame, text="Statistics")
        stats_frame.pack(fill=tk.X, pady=10)

        # Total eggs
        ttk.Label(stats_frame, text="Total Eggs Counted:").grid(
            row=0, column=0, sticky=tk.W, pady=2)
        self.total_eggs_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.total_eggs_var, font=(
            'Arial', 12, 'bold')).grid(row=0, column=1, sticky=tk.W, pady=2)

        # Current eggs
        ttk.Label(stats_frame, text="Current Eggs:").grid(
            row=1, column=0, sticky=tk.W, pady=2)
        self.current_eggs_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.current_eggs_var).grid(
            row=1, column=1, sticky=tk.W, pady=2)

        # Tracked objects
        ttk.Label(stats_frame, text="Tracked Objects:").grid(
            row=2, column=0, sticky=tk.W, pady=2)
        self.tracked_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.tracked_var).grid(
            row=2, column=1, sticky=tk.W, pady=2)

        # FPS
        ttk.Label(stats_frame, text="FPS:").grid(
            row=3, column=0, sticky=tk.W, pady=2)
        self.fps_var = tk.StringVar(value="0.0")
        ttk.Label(stats_frame, textvariable=self.fps_var).grid(
            row=3, column=1, sticky=tk.W, pady=2)

        # Classification Stats
        ttk.Label(stats_frame, text="White Egg:").grid(
            row=4, column=0, sticky=tk.W, pady=2)
        self.white_egg_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.white_egg_var).grid(
            row=4, column=1, sticky=tk.W, pady=2)

        ttk.Label(stats_frame, text="Dirty Egg:").grid(
            row=5, column=0, sticky=tk.W, pady=2)
        self.dirty_egg_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.dirty_egg_var).grid(
            row=5, column=1, sticky=tk.W, pady=2)

        ttk.Label(stats_frame, text="Broken Egg:").grid(
            row=6, column=0, sticky=tk.W, pady=2)
        self.broken_egg_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.broken_egg_var).grid(
            row=6, column=1, sticky=tk.W, pady=2)

        # Pause status indicator
        ttk.Label(stats_frame, text="Counting Status:").grid(
            row=7, column=0, sticky=tk.W, pady=2)
        self.counting_status_var = tk.StringVar(value="Active")
        self.counting_status_label = ttk.Label(
            stats_frame, textvariable=self.counting_status_var)
        self.counting_status_label.grid(row=7, column=1, sticky=tk.W, pady=2)
        self.update_counting_status_display()

        # Role indicator
        ttk.Label(stats_frame, text="User Role:").grid(
            row=8, column=0, sticky=tk.W, pady=2)
        self.role_var = tk.StringVar(value=self.role.capitalize())
        role_label = ttk.Label(
            stats_frame, textvariable=self.role_var, font=('Arial', 10, 'bold'))
        role_label.grid(row=8, column=1, sticky=tk.W, pady=2)
        role_label.config(foreground='green' if self.is_admin else 'blue')

        # Export status
        self.export_status_var = tk.StringVar(value="Excel Export: Ready")
        ttk.Label(scrollable_frame, textvariable=self.export_status_var,
                  font=('Arial', 9)).pack(pady=5)

        # Stream status
        self.stream_status_var = tk.StringVar(value="Stream: Stopped")
        ttk.Label(scrollable_frame, textvariable=self.stream_status_var,
                  font=('Arial', 9)).pack(pady=5)

        stats_frame.columnconfigure(1, weight=1)

        # NEW: Controls before logout section
        controls_before_logout_frame = ttk.LabelFrame(
            scrollable_frame, text="Quick Actions")
        controls_before_logout_frame.pack(fill=tk.X, pady=(10, 5))

        # Create a frame for the quick action buttons
        quick_actions_frame = ttk.Frame(controls_before_logout_frame)
        quick_actions_frame.pack(pady=5)

        # Create a container frame
        controls_container = ttk.Frame(quick_actions_frame)
        controls_container.pack(fill=tk.X, expand=True)

        # Top row for Pause and Hide buttons
        top_row = ttk.Frame(controls_container)
        top_row.pack(fill=tk.X, pady=(0, 10))

        self.pause_button = ttk.Button(top_row, text="Pause Counting",
                                       command=self.toggle_pause_counting, width=20)
        self.pause_button.pack(side=tk.LEFT, padx=(5, 10))

        ttk.Button(top_row, text="Hide Controls",
                   command=self.toggle_controls_visibility, width=20).pack(side=tk.LEFT, padx=(10, 5))

        # Bottom row for Enable Recording
        bottom_row = ttk.Frame(controls_container)
        bottom_row.pack(fill=tk.X)

        self.record_var = tk.BooleanVar(value=False)
        record_cb = ttk.Checkbutton(bottom_row, text="Enable Recording",
                                    variable=self.record_var)
        record_cb.pack(side=tk.LEFT, padx=(5, 0))

    def build_operator_controls_tab(self, parent):
        """Build controls tab for operator with ALL features enabled"""

        # Logout button frame (Fixed at bottom)
        logout_frame = ttk.Frame(parent)
        logout_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10, padx=10)

        logout_btn = ttk.Button(logout_frame, text="⏻ Logout",
                                command=self.logout, style='Danger.TButton', width=20)
        logout_btn.pack(side=tk.BOTTOM)

        # Scrollable container for controls
        canvas = tk.Canvas(
            parent, bg=THEME_COLORS['bg_dark'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(
            parent, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Helper for sections
        def create_section(title):
            frame = ttk.LabelFrame(scrollable_frame, text=title, padding=10)
            frame.pack(fill=tk.X, padx=10, pady=5)
            return frame

        # Main controls
        controls_frame = create_section("System Control")

        self.start_btn = ttk.Button(controls_frame, text="Start System",
                                    command=self.start_system, width=15, style='Accent.TButton')
        self.start_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.stop_btn = ttk.Button(controls_frame, text="⏹ Stop System",
                                   command=self.stop_system, width=15)
        self.stop_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.load_model_btn = ttk.Button(controls_frame, text="📂 Load Model",
                                         command=self.load_model, width=15)
        # Hidden for operator - combined with Start System
        # # self.load_model_btn.pack(side=tk.LEFT) # Hidden for operator

        # Counting controls
        counting_frame = create_section("Counting Operations")

        self.reset_btn = ttk.Button(counting_frame, text="↺ Reset Count",
                                    command=self.reset_count, width=15)
        self.reset_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.adjust_btn = ttk.Button(counting_frame, text="📏 Adjust Divider",
                                     command=self.toggle_adjustment, width=15)
        self.adjust_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.capture_btn = ttk.Button(counting_frame, text="📷 Capture Image",
                                      command=self.capture_image, width=15)
        self.capture_btn.pack(side=tk.LEFT)

        # Performance controls - ENABLED FOR OPERATORS
        perf_frame = create_section("Performance & Display")

        self.toggle_device_btn = ttk.Button(perf_frame, text="⚡ Toggle GPU/CPU",
                                            command=self.toggle_device, width=15)
        self.toggle_device_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.save_settings_btn = ttk.Button(perf_frame, text="💾 Save Settings",
                                            command=self.save_settings, width=15)
        self.save_settings_btn.pack(side=tk.LEFT)

        # Full screen toggle button
        ttk.Button(perf_frame, text="⛶ Toggle Full Screen ",
                   command=self.toggle_full_screen, width=20).pack(side=tk.LEFT, padx=(10, 5))

        # Excel Export Controls - ENABLED FOR OPERATORS
        excel_frame = create_section("Data Export")

        self.open_excel_btn = ttk.Button(excel_frame, text="📊 Open Excel",
                                         command=self.open_excel_file, width=12)
        self.open_excel_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.daily_report_btn = ttk.Button(excel_frame, text="📅 Daily Report",
                                           command=self.create_daily_report, width=12)
        self.daily_report_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.export_status_btn = ttk.Button(excel_frame, text="ℹ Status",
                                            command=self.show_export_status, width=12)
        self.export_status_btn.pack(side=tk.LEFT)

        # Streaming Controls - ENABLED FOR OPERATORS
        stream_frame = create_section("Wireless Streaming")

        self.start_stream_btn = ttk.Button(stream_frame, text="📡 Start Stream",
                                           command=self.start_streaming, width=15)
        self.start_stream_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.stop_stream_btn = ttk.Button(stream_frame, text="⏹ Stop Stream",
                                          command=self.stop_streaming, width=15)
        self.stop_stream_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.stream_info_btn = ttk.Button(stream_frame, text="ℹ Info",
                                          command=self.show_stream_info, width=12)
        self.stream_info_btn.pack(side=tk.LEFT)

        # Statistics display
        stats_frame = ttk.LabelFrame(scrollable_frame, text="Statistics")
        stats_frame.pack(fill=tk.X, pady=10)

        # Total eggs
        ttk.Label(stats_frame, text="Total Eggs Counted:").grid(
            row=0, column=0, sticky=tk.W, pady=2)
        self.total_eggs_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.total_eggs_var, font=(
            'Arial', 12, 'bold')).grid(row=0, column=1, sticky=tk.W, pady=2)

        # Current eggs
        ttk.Label(stats_frame, text="Current Eggs:").grid(
            row=1, column=0, sticky=tk.W, pady=2)
        self.current_eggs_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.current_eggs_var).grid(
            row=1, column=1, sticky=tk.W, pady=2)

        # Tracked objects
        ttk.Label(stats_frame, text="Tracked Objects:").grid(
            row=2, column=0, sticky=tk.W, pady=2)
        self.tracked_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.tracked_var).grid(
            row=2, column=1, sticky=tk.W, pady=2)

        # FPS
        ttk.Label(stats_frame, text="FPS:").grid(
            row=3, column=0, sticky=tk.W, pady=2)
        self.fps_var = tk.StringVar(value="0.0")
        ttk.Label(stats_frame, textvariable=self.fps_var).grid(
            row=3, column=1, sticky=tk.W, pady=2)

        # Classification Stats
        ttk.Label(stats_frame, text="White Egg:").grid(
            row=4, column=0, sticky=tk.W, pady=2)
        # Re-use the same variables as admin tab if feasible, but here we are re-building.
        # Since variables are bound to self, rebuilding overwrites them or re-binds them.
        # But wait, self.white_egg_var might have been created in build_controls_tab if that was called first.
        # However, usually only one is built. 
        # But to be safe and consistent, we re-declare or re-use. 
        # If we re-declare, the update loop needs to update THIS instance's var.
        # But self.white_egg_var is an instance attribute. 
        # So defining it here again overwrites the reference, which is fine as long as 
        # update_counting_status_display uses self.white_egg_var.
        
        # To avoid issues if switching tabs (though this app seems to rebuild or toggle frames),
        # let's just assign.
        self.white_egg_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.white_egg_var).grid(
            row=4, column=1, sticky=tk.W, pady=2)

        ttk.Label(stats_frame, text="Dirty Egg:").grid(
            row=5, column=0, sticky=tk.W, pady=2)
        self.dirty_egg_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.dirty_egg_var).grid(
            row=5, column=1, sticky=tk.W, pady=2)

        ttk.Label(stats_frame, text="Broken Egg:").grid(
            row=6, column=0, sticky=tk.W, pady=2)
        self.broken_egg_var = tk.StringVar(value="0")
        ttk.Label(stats_frame, textvariable=self.broken_egg_var).grid(
            row=6, column=1, sticky=tk.W, pady=2)

        # Pause status indicator
        ttk.Label(stats_frame, text="Counting Status:").grid(
            row=7, column=0, sticky=tk.W, pady=2)
        self.counting_status_var = tk.StringVar(value="Active")
        self.counting_status_label = ttk.Label(
            stats_frame, textvariable=self.counting_status_var)
        self.counting_status_label.grid(row=7, column=1, sticky=tk.W, pady=2)
        self.update_counting_status_display()

        # Role indicator
        ttk.Label(stats_frame, text="User Role:").grid(
            row=8, column=0, sticky=tk.W, pady=2)
        self.role_var = tk.StringVar(value=self.role.capitalize())
        role_label = ttk.Label(
            stats_frame, textvariable=self.role_var, font=('Arial', 10, 'bold'))
        role_label.grid(row=8, column=1, sticky=tk.W, pady=2)
        role_label.config(foreground='blue')

        # Export status
        self.export_status_var = tk.StringVar(value="Excel Export: Ready")
        ttk.Label(scrollable_frame, textvariable=self.export_status_var,
                  font=('Arial', 9)).pack(pady=5)

        # Stream status
        self.stream_status_var = tk.StringVar(value="Stream: Stopped")
        ttk.Label(scrollable_frame, textvariable=self.stream_status_var,
                  font=('Arial', 9)).pack(pady=5)

        stats_frame.columnconfigure(1, weight=1)

        # NEW: Controls before logout section for operators
        controls_before_logout_frame = ttk.LabelFrame(
            scrollable_frame, text="Quick Actions")
        controls_before_logout_frame.pack(fill=tk.X, pady=(10, 5))

        # Create a frame for the quick action buttons
        quick_actions_frame = ttk.Frame(controls_before_logout_frame)
        quick_actions_frame.pack(pady=5)

        # Create a container frame
        controls_container = ttk.Frame(quick_actions_frame)
        controls_container.pack(fill=tk.X, expand=True)

        # Top row for Pause and Hide buttons
        top_row = ttk.Frame(controls_container)
        top_row.pack(fill=tk.X, pady=(0, 10))

        self.pause_button = ttk.Button(top_row, text="Pause Counting",
                                       command=self.toggle_pause_counting, width=20)
        self.pause_button.pack(side=tk.LEFT, padx=(5, 10))

        ttk.Button(top_row, text="Hide Controls",
                   command=self.toggle_controls_visibility, width=20).pack(side=tk.LEFT, padx=(10, 5))

        # Bottom row for Enable Recording
        bottom_row = ttk.Frame(controls_container)
        bottom_row.pack(fill=tk.X)

        self.record_var = tk.BooleanVar(value=False)
        record_cb = ttk.Checkbutton(bottom_row, text="Enable Recording",
                                    variable=self.record_var)
        record_cb.pack(side=tk.LEFT, padx=(5, 0))

        # Quick Settings Section for Operators
        settings_frame = ttk.LabelFrame(
            scrollable_frame, text="Quick Settings")
        settings_frame.pack(fill=tk.X, pady=10)

    def build_video_display(self, parent):
        """Build video display area"""
        # Video frame
        video_frame = ttk.LabelFrame(parent, text="Live Feed", padding=5)
        video_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # Canvas for video - darker background for video area
        self.video_canvas = tk.Canvas(
            video_frame, bg='black', width=800, height=600, highlightthickness=0)
        self.video_canvas.pack(fill=tk.BOTH, expand=True)

        # Log frame
        self.log_frame = ttk.LabelFrame(parent, text="System Log", padding=5)
        # Remove bottom padding logic handled by parent
        self.log_frame.pack(fill=tk.BOTH, expand=False, pady=(0, 0))

        # Text widget for logs - Custom colors for dark mode context
        self.log_text = scrolledtext.ScrolledText(
            self.log_frame, height=8, width=80,
            bg=THEME_COLORS['bg_lighter'], fg=THEME_COLORS['text_main'],
            insertbackground='white', relief='flat', font=('Consolas', 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)
        self.log_text.config(state=tk.DISABLED)

    def build_status_bar_placeholder(self):
        pass  # Moved to separate method to override properly

    def setup_bindings(self):
        """Setup keyboard bindings"""
        self.root.bind('<KeyPress>', self.handle_keypress)
        # F11 for full screen toggle
        self.root.bind('<F11>', self.toggle_full_screen)
        self.video_canvas.bind('<Button-1>', self.handle_canvas_click)
        self.video_canvas.bind('<B1-Motion>', self.handle_canvas_drag)
        self.video_canvas.bind('<ButtonRelease-1>', self.handle_canvas_release)

    def on_closing(self):
        """Handle window closing event safely"""
        if messagebox.askokcancel("Quit", "Do you want to stop the system and quit?"):
            self.stop_system()
            self.root.destroy()

    def logout(self):
        """Logout and return to login screen"""
        role_text = f"({self.role.capitalize()} account)"
        if messagebox.askyesno("Logout", f"Are you sure you want to logout? {role_text}"):
            self.stop_system()

            # Clear current window
            for widget in self.root.winfo_children():
                widget.destroy()

            # Re-initialize login screen
            LoginApp(self.root)

    def toggle_controls_visibility(self):
        """Toggle visibility of controls (left panel and log)"""
        self.controls_visible = not self.controls_visible

        if self.controls_visible:
            # Show controls
            self.left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5))
            self.log_frame.pack(fill=tk.BOTH, expand=False, pady=(0, 10))
            self.log_message("Controls shown - Press 'H' to hide")
        else:
            # Hide controls
            self.left_frame.pack_forget()
            self.log_frame.pack_forget()
            self.log_message("Controls hidden - Press 'H' to show")

    def toggle_full_screen(self, event=None):
        """Toggle between full screen and normal mode"""
        self.full_screen_mode = not self.full_screen_mode

        if self.full_screen_mode:
            # Enter full screen mode
            self.root.attributes('-fullscreen', True)
            self.root.geometry("{0}x{1}+0+0".format(
                self.root.winfo_screenwidth(), self.root.winfo_screenheight()))

            # Hide controls when entering full screen
            self.controls_visible = False
            self.left_frame.pack_forget()
            self.log_frame.pack_forget()
            self.status_frame.pack_forget()

            # Make video canvas fill the entire window
            self.video_canvas.pack_forget()
            self.video_canvas.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

            self.log_message("Full screen mode activated")
            self.log_message(
                "Press 'H' to show controls, 'F11' to exit full screen")
        else:
            # Exit full screen mode
            self.root.attributes('-fullscreen', False)

            # Restore normal layout
            self.video_canvas.pack_forget()
            self.main_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

            # Show controls if they were visible before
            if self.controls_visible:
                self.left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5))

            self.right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
            self.video_canvas.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

            if self.controls_visible:
                self.log_frame.pack(fill=tk.BOTH, expand=False, pady=(0, 10))

            self.status_frame.pack(fill=tk.X, side=tk.BOTTOM)

            # Restore window size
            self.root.geometry("1400x900")

            self.log_message("Normal mode activated")

    def log_message(self, message):
        """Add message to log"""
        # Safety check for early calls before UI is ready
        if not hasattr(self, 'log_text') or self.log_text is None:
            print(f"[Waiting for UI] {message}")
            return

        try:
            self.log_text.config(state=tk.NORMAL)
            timestamp = datetime.datetime.now().strftime("%H:%M:%S")
            self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
            self.log_text.see(tk.END)
            self.log_text.config(state=tk.DISABLED)
        except Exception as e:
            # Fallback if widget is destroyed/invalid
            print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}")

    def browse_model(self):
        """Browse for YOLO model file"""
        filename = filedialog.askopenfilename(
            title="Select YOLO Model",
            filetypes=[("YOLO Model", "*.pt"), ("All files", "*.*")]
        )
        if filename:
            self.model_path_var.set(filename)
            self.settings.model_path = filename
            self.save_settings()

    def browse_source(self):
        """Browse for source file or folder based on selection"""
        source_type = self.source_var.get()

        if source_type == 'Image File':
            filename = filedialog.askopenfilename(
                title="Select Image File",
                filetypes=[
                    ("Image files", "*.jpg *.jpeg *.png *.bmp *.tiff"), ("All files", "*.*")]
            )
            if filename:
                self.source_var.set(filename)
                self.settings.source = filename
                self.save_settings()

        elif source_type == 'Image Folder':
            folder = filedialog.askdirectory(title="Select Image Folder")
            if folder:
                self.source_var.set(folder)
                self.settings.source = folder
                self.save_settings()

        elif source_type == 'Video File':
            filename = filedialog.askopenfilename(
                title="Select Video File",
                filetypes=[
                    ("Video files", "*.mp4 *.avi *.mov *.mkv *.wmv"), ("All files", "*.*")]
            )
            if filename:
                self.source_var.set(filename)
                self.settings.source = filename
                self.save_settings()

    def connect_plc(self):
        """Connect to PLC"""
        try:
            self.plc = ModbusTCPPLC(
                ip=self.plc_ip_var.get() if hasattr(self, 'plc_ip_var') else self.settings.plc_ip,
                port=self.plc_port_var.get() if hasattr(
                    self, 'plc_port_var') else self.settings.plc_port,
                unit_id=self.plc_unit_id_var.get() if hasattr(
                    self, 'plc_unit_id_var') else self.settings.plc_unit_id
            )
            if self.plc.connect():
                self.plc_status_var.set("Connected")
                self.log_message(
                    f"PLC connected to {self.plc.ip}:{self.plc.port}")
            else:
                self.plc_status_var.set("Connection Failed")
                self.log_message("PLC connection failed")
        except Exception as e:
            self.log_message(f"PLC connection error: {e}")

    def disconnect_plc(self):
        """Disconnect from PLC"""
        if self.plc and hasattr(self.plc, 'connected') and self.plc.connected:
            self.plc.disconnect()
            self.plc_status_var.set("Disconnected")
            self.log_message("PLC disconnected")

    def test_plc(self):
        """Test PLC communication"""
        if self.plc and hasattr(self.plc, 'connected') and self.plc.connected:
            if self.plc.test_communication():
                self.log_message("PLC communication test successful")
            else:
                self.log_message("PLC communication test failed")
        else:
            self.log_message("PLC not connected")

    def load_model(self):
        """Load YOLO model in background thread with better device handling"""
        if self.running:
            self.log_message("⚠️ Cannot load model while system is running. Stop the system first.")
            return

        if self.loading_model:
            return

        self.loading_model = True
        # Get model path from settings for operators, or from GUI for admins
        if self.is_admin and hasattr(self, 'model_path_var'):
            model_path = self.model_path_var.get()
        else:
            model_path = self.settings.model_path

        if not model_path or model_path == '':
            messagebox.showerror("Error", "Please select a model file first")
            return

        # Get device setting
        if self.is_admin and hasattr(self, 'device_var'):
            requested_device = self.device_var.get()
        else:
            requested_device = self.settings.device

        # Validate device selection
        if requested_device.startswith('cuda'):
            if not torch.cuda.is_available():
                self.log_message(
                    "⚠️ CUDA requested but not available. Switching to CPU.")
                requested_device = 'cpu'
                if hasattr(self, 'device_var'):
                    self.device_var.set('cpu')

        # Disable buttons while loading
        self._update_button_states()
        self.load_model_btn.config(text="Loading...")

        def load_model_thread():
            try:
                self.log_message(
                    f"Loading YOLO model on {requested_device.upper()}...")

                # Get half precision setting
                if self.is_admin and hasattr(self, 'half_precision_var'):
                    half_precision = self.half_precision_var.get()
                else:
                    half_precision = self.settings.use_half_precision

                success = self.inference_engine.load_model(
                    model_path=model_path,
                    device=requested_device,
                    half_precision=half_precision
                )

                if success:
                    self.model_loaded = True
                    self.device_status_var.set(
                        f"Device: {self.inference_engine.device.upper()}")
                    self.log_message(
                        f"Model loaded successfully on {self.inference_engine.device.upper()}!")
                else:
                    self.log_message("Failed to load model")
                    self.model_loaded = False

            except Exception as e:
                self.log_message(f"Error loading model: {e}")
                self.model_loaded = False
            finally:
                self.loading_model = False
                self.root.after(0, self._update_button_states)
                self.root.after(0, lambda: self.load_model_btn.config(text="Load Model"))

        threading.Thread(target=load_model_thread, daemon=True).start()

    def toggle_device(self):
        """Toggle between GPU and CPU with better error handling"""
        try:
            # Get current device
            current_device = self.device_var.get() if hasattr(
                self, 'device_var') else self.settings.device

            # Check CUDA availability in detail
            cuda_available = torch.cuda.is_available()

            if not cuda_available:
                # Provide detailed information about why GPU is not available
                self.log_message("⚠️ GPU NOT AVAILABLE - Detailed Check:")
                self.log_message(
                    f"  • PyTorch CUDA available: {torch.cuda.is_available()}")
                self.log_message(f"  • PyTorch version: {torch.__version__}")

                try:
                    # Check CUDA version
                    if hasattr(torch.version, 'cuda'):
                        self.log_message(
                            f"  • PyTorch CUDA version: {torch.version.cuda}")
                    else:
                        self.log_message(
                            "  • PyTorch compiled without CUDA support")
                except:
                    self.log_message("  • Could not determine CUDA version")

                # If trying to switch to GPU but not available, stay on CPU
                if current_device.startswith('cuda'):
                    new_device = 'cpu'
                    self.log_message(f"  • Forcing CPU mode (GPU unavailable)")
                else:
                    new_device = 'cpu'
                    self.log_message(f"  • Already using CPU mode")

            else:
                # CUDA is available
                gpu_name = torch.cuda.get_device_name(
                    0) if torch.cuda.device_count() > 0 else "Unknown GPU"
                self.log_message(f"✅ GPU Available: {gpu_name}")
                self.log_message(
                    f"  • CUDA Devices: {torch.cuda.device_count()}")
                self.log_message(
                    f"  • Current PyTorch CUDA version: {torch.version.cuda}")

                # Toggle between CPU and GPU
                if current_device.startswith('cuda'):
                    new_device = 'cpu'
                    self.log_message(f"  • Switching from GPU to CPU")
                else:
                    new_device = 'cuda:0'
                    self.log_message(f"  • Switching from CPU to GPU")

            # Update the device setting
            if hasattr(self, 'device_var'):
                self.device_var.set(new_device)
            else:
                setattr(self.settings, 'device', new_device)

            self.log_message(f"Device set to: {new_device.upper()}")

            # Reload model with new device if model is loaded
            if self.inference_engine.is_loaded:
                self.inference_engine.is_loaded = False
                self.device_status_var.set(
                    f"Switching to {new_device.upper()}...")
                self.load_model()
            else:
                self.device_status_var.set(
                    f"Device: {new_device.upper()} (Model not loaded)")

        except Exception as e:
            self.log_message(f"❌ Error toggling device: {str(e)}")
            # Fall back to CPU
            if hasattr(self, 'device_var'):
                self.device_var.set('cpu')
            else:
                setattr(self.settings, 'device', 'cpu')
            self.log_message("⚠️ Fallback to CPU mode due to error")

    def start_system(self):
        """Start the egg counting system with optimizations"""
        if self.running:
            return

        if self.loading_model:
            self.log_message("⚠️ System is currently loading a model. Please wait...")
            return

        if not self.inference_engine.is_loaded:
            if not self.is_admin:
                self.log_message("Model not loaded. Loading model automatically...")
                self.load_model()
                self._wait_for_model_and_start()
                return
            else:
                messagebox.showwarning("Warning", "Please load model first")
                return

        # Update settings from GUI
        self._update_settings_from_gui()

        # Ensure tracker starts with persistent count if it's currently 0 or smaller than stored
        if self.egg_tracker.total_eggs_counted == 0:
            self.egg_tracker.set_initial_count(self.settings.persistent_count)

        # Initialize source
        if not self._initialize_source():
            return

        # Connect to PLC (for both admin and operator)
        if self.settings.plc_ip:
            self.connect_plc()

        # Start data exporters
        self.csv_exporter.start()
        self.excel_exporter.start()

        # Start streaming (for both admin and operator)
        self.start_streaming()

        self.running = True
        self._update_button_states()
        self.system_status_var.set(
            f"System Running - {self.role.capitalize()} Mode")
        self.log_message("Egg counting system started")
        self.log_message("Excel/CSV export started (updates every minute)")
        self.log_message("Wireless streaming available on port 8080")
        self.log_message(
            "Press SPACEBAR to pause/resume counting during maintenance")
        self.log_message("Press 'H' to hide/show controls")

        # Start video processing in separate thread
        self.video_thread = threading.Thread(
            target=self._video_processing_loop, daemon=True)
        self.video_thread.start()

        # Start display update in main thread
        self._start_display_update()

    def start_streaming(self):
        """Start wireless streaming"""
        try:
            # Revert to hardcoded port 8080 as requested
            self.mjpeg_streamer.port = 8080
            
            success = self.mjpeg_streamer.start()
            if success:
                self.stream_status_var.set("Stream: Active on port 8080")
                local_ip = self.mjpeg_streamer._get_local_ip()
                self.log_message(
                    f"Access stream from any device: http://{local_ip}:8080")
            else:
                self.stream_status_var.set("Stream: Failed to start")
        except Exception as e:
            self.log_message(f"Stream start error: {e}")

    def stop_streaming(self):
        """Stop wireless streaming"""
        self.mjpeg_streamer.stop()
        self.stream_status_var.set("Stream: Stopped")
        self.log_message("Wireless streaming stopped")

    def _update_settings_from_gui(self):
        """Update settings from GUI values"""
        # Update settings for both admin and operator
        if hasattr(self, 'model_path_var'):
            self.settings.model_path = self.model_path_var.get()
        if hasattr(self, 'source_var'):
            self.settings.source = self.source_var.get()
        if hasattr(self, 'threshold_var'):
            self.settings.threshold = self.threshold_var.get()
        if hasattr(self, 'resolution_var'):
            self.settings.resolution = self.resolution_var.get()
        if hasattr(self, 'device_var'):
            self.settings.device = self.device_var.get()
        if hasattr(self, 'half_precision_var'):
            self.settings.use_half_precision = self.half_precision_var.get()
        if hasattr(self, 'plc_ip_var'):
            self.settings.plc_ip = self.plc_ip_var.get()
            self.settings.plc_port = self.plc_port_var.get()
            self.settings.plc_address = self.plc_address_var.get()
            self.settings.plc_unit_id = self.plc_unit_id_var.get()
        if hasattr(self, 'rotate_var'):
            self.settings.rotate = self.rotate_var.get()
        if hasattr(self, 'divider_upper_var'):
            self.settings.divider_ratio_upper = self.divider_upper_var.get()
        if hasattr(self, 'divider_lower_var'):
            self.settings.divider_ratio_lower = self.divider_lower_var.get()
        if hasattr(self, 'frame_skip_var'):
            self.settings.frame_skip = self.frame_skip_var.get()
        if hasattr(self, 'imgsz_var'):
            self.settings.inference_imgsz = self.imgsz_var.get()
        if hasattr(self, 'show_overlays_var'):
            self.settings.show_overlays = self.show_overlays_var.get()

        # Update tracker settings
        if hasattr(self, 'max_distance_var'):
            self.egg_tracker.max_distance = self.max_distance_var.get()
        if hasattr(self, 'plc_update_interval_var'):
            self.plc_update_interval = self.plc_update_interval_var.get()

        self.save_settings()

    def _initialize_source(self):
        """Initialize source based on selection"""
        if hasattr(self, 'source_var'):
            source = self.source_var.get()
        else:
            source = self.settings.source

        try:
            # Determine source type
            if source in ['0', '1', '2']:  # Webcam indices
                self.current_source_type = 'camera'
                return self._initialize_camera()
            elif source == 'deltacam':
                self.current_source_type = 'camera'
                return self._initialize_delta_camera()
            elif source.endswith(('.jpg', '.jpeg', '.png', '.bmp', '.tiff')):
                self.current_source_type = 'image'
                return self._initialize_image_file()
            elif source.endswith(('.mp4', '.avi', '.mov', '.mkv', '.wmv')):
                self.current_source_type = 'video'
                return self._initialize_video_file()
            elif os.path.isdir(source):
                self.current_source_type = 'folder'
                return self._initialize_image_folder()
            else:
                self.log_message(f"Unknown source type: {source}")
                return False

        except Exception as e:
            self.log_message(f"Source initialization error: {e}")
            return False

    def _wait_for_model_and_start(self):
        """Wait for the model to load before starting the system (for operator)"""
        if self.model_loaded:
            self.log_message("Model loaded. Proceeding to start system...")
            self.start_system()
        elif self.load_model_btn.cget('state') == 'normal':
            # This means loading finished (possibly failed if model_loaded is still False)
            if not self.model_loaded:
                self.log_message("❌ Model loading failed. System could not start.")
        else:
            # Still loading (button is disabled), check again in 500ms
            self.root.after(500, self._wait_for_model_and_start)

    def _update_button_states(self):
        """Update button enabled/disabled states based on system status"""
        is_running = self.running
        is_loading = self.loading_model

        if hasattr(self, 'start_btn'):
            self.start_btn.config(state='disabled' if is_running or is_loading else 'normal')
        if hasattr(self, 'stop_btn'):
            self.stop_btn.config(state='normal' if is_running else 'disabled')
        if hasattr(self, 'load_model_btn'):
            self.load_model_btn.config(state='disabled' if is_running or is_loading else 'normal')
        if hasattr(self, 'source_combo'): # Check if we have a reference to combos
            pass # We could disable more inputs here if needed

    def _initialize_camera(self):
        """Initialize web camera source with performance optimizations"""
        if hasattr(self, 'source_var'):
            source = self.source_var.get()
        else:
            source = self.settings.source

        try:
            camera_index = int(source)
            # Try DirectShow (DSHOW) first on Windows as it's often more stable for webcams
            self.cap = cv2.VideoCapture(camera_index, cv2.CAP_DSHOW)
            if not self.cap.isOpened():
                # Fallback to default backend if DSHOW fails
                self.cap = cv2.VideoCapture(camera_index)
            
            # Stabilization fix: Disable multi-threaded decoding within the capture object
            # Use getattr for compatibility with older OpenCV versions (Attribute ID 45)
            self.cap.set(getattr(cv2, 'CAP_PROP_THREAD_COUNT', 45), 1)

            # Performance optimizations for web cameras
            # Reduce buffer to minimize latency
            self.cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
            self.cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(
                'M', 'J', 'P', 'G'))  # Use MJPEG codec
            self.cap.set(cv2.CAP_PROP_FPS, 30)  # Set explicit FPS

            # Set resolution if specified
            if hasattr(self, 'resolution_var') and self.resolution_var.get():
                resW, resH = map(int, self.resolution_var.get().split('x'))
                self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, resW)
                self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, resH)
            else:
                # Use default resolution
                self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
                self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

            # Auto-exposure and other optimizations
            self.cap.set(cv2.CAP_PROP_AUTO_EXPOSURE, 1)  # Enable auto-exposure
            # Disable autofocus for better performance
            self.cap.set(cv2.CAP_PROP_AUTOFOCUS, 0)

            # Test camera with timeout
            import time
            start_time = time.time()
            while time.time() - start_time < 5:  # 5 second timeout
                ret, test_frame = self.cap.read()
                if ret:
                    break
                time.sleep(0.1)

            if not ret:
                self.log_message("ERROR: Unable to read from web camera")
                self.cap.release()
                self.cap = None
                return False

            self.camera_status_var.set("Camera: Webcam Connected")
            self.log_message(
                f"Web camera {camera_index} initialized with performance settings")
            return True

        except Exception as e:
            self.log_message(f"Camera initialization error: {e}")
            if hasattr(self, 'cap') and self.cap:
                self.cap.release()
                self.cap = None
            return False

    def _initialize_delta_camera(self):
        """Initialize Delta camera source"""
        try:
            self.delta_cam = DeltaCamera()
            if not self.delta_cam.connect():
                self.log_message("ERROR: Failed to connect to Delta camera")
                return False

            # Apply Delta camera specific optimizations
            self.delta_cam.start_acquisition()

            # Set Delta camera to performance mode if available
            if hasattr(self.delta_cam, 'set_performance_mode'):
                self.delta_cam.set_performance_mode(True)

            # Test with frame skipping for initialization
            for _ in range(5):  # Skip first few frames which might be buffered
                ret, test_frame = self.delta_cam.read_frame()
                if not ret:
                    self.log_message("ERROR: Unable to read from Delta camera")
                    return False

            self.camera_status_var.set("Camera: Delta Connected")
            self.log_message(
                "Delta camera initialized with performance settings")
            return True

        except Exception as e:
            self.log_message(f"Delta camera initialization error: {e}")
            return False

    def _initialize_image_file(self):
        """Initialize single image file source"""
        if hasattr(self, 'source_var'):
            image_path = self.source_var.get()
        else:
            image_path = self.settings.source

        if not os.path.exists(image_path):
            self.log_message(f"Image file not found: {image_path}")
            return False

        self.image_files = [image_path]
        self.current_image_index = 0
        self.camera_status_var.set("Source: Image File")
        self.log_message(f"Image file loaded: {os.path.basename(image_path)}")
        return True

    def _initialize_image_folder(self):
        """Initialize image folder source"""
        if hasattr(self, 'source_var'):
            folder_path = self.source_var.get()
        else:
            folder_path = self.settings.source

        if not os.path.exists(folder_path):
            self.log_message(f"Folder not found: {folder_path}")
            return False

        # Get all image files from folder
        extensions = ['*.jpg', '*.jpeg', '*.png', '*.bmp', '*.tiff']
        self.image_files = []
        for ext in extensions:
            self.image_files.extend(glob.glob(os.path.join(folder_path, ext)))
            self.image_files.extend(
                glob.glob(os.path.join(folder_path, ext.upper())))

        if not self.image_files:
            self.log_message(f"No image files found in folder: {folder_path}")
            return False

        self.image_files.sort()
        self.current_image_index = 0
        self.camera_status_var.set(
            f"Source: Image Folder ({len(self.image_files)} images)")
        self.log_message(f"Loaded {len(self.image_files)} images from folder")
        return True

    def _initialize_video_file(self):
        """Initialize video file source"""
        if hasattr(self, 'source_var'):
            video_path = self.source_var.get()
        else:
            video_path = self.settings.source

        if not os.path.exists(video_path):
            self.log_message(f"Video file not found: {video_path}")
            return False

        # Open with explicit thread count limit to fix FFmpeg assertion errors
        # Use getattr for compatibility with older OpenCV versions (Attribute ID 45)
        self.cap = cv2.VideoCapture(video_path)
        self.cap.set(getattr(cv2, 'CAP_PROP_THREAD_COUNT', 45), 1)

        if not self.cap.isOpened():
            self.log_message(f"Failed to open video file: {video_path}")
            return False

        self.camera_status_var.set("Source: Video File")
        self.log_message(f"Video file loaded: {os.path.basename(video_path)}")
        return True

    def _video_processing_loop(self):
        """Optimized video processing loop"""
        while self.running:
            t_start = time.perf_counter()

            # Read frame based on source type
            frame = self._read_frame_from_source()
            if frame is None:
                time.sleep(0.01)
                continue

            # Process frame
            processed_frame = self._process_frame(frame)
            if processed_frame is None:
                continue

            # Update display queue (non-blocking)
            if not self.frame_queue.full():
                try:
                    self.frame_queue.put_nowait(processed_frame)
                except queue.Full:
                    pass  # Skip frame if queue is full

            # Calculate FPS
            t_stop = time.perf_counter()
            frame_rate = 1.0 / (t_stop - t_start)

            # Update FPS buffer efficiently
            self.frame_rate_buffer.append(frame_rate)
            self.avg_frame_rate = np.mean(self.frame_rate_buffer)

            self.frame_counter += 1

    def _read_frame_from_source(self):
        """Read frame based on current source type"""
        try:
            if self.current_source_type == 'camera':
                if self.settings.source == 'deltacam' and self.delta_cam and self.delta_cam.connected:
                    ret, frame = self.delta_cam.read_frame()
                    return frame if ret else None
                elif self.cap and self.cap.isOpened():
                    ret, frame = self.cap.read()
                    return frame if ret else None

            elif self.current_source_type == 'image':
                if self.current_image_index < len(self.image_files):
                    frame = cv2.imread(
                        self.image_files[self.current_image_index])
                    # Auto-advance to next image after a delay for demonstration
                    time.sleep(0.1)  # Simulate frame rate for images
                    self.current_image_index = (
                        self.current_image_index + 1) % len(self.image_files)
                    return frame

            elif self.current_source_type == 'folder':
                if self.current_image_index < len(self.image_files):
                    frame = cv2.imread(
                        self.image_files[self.current_image_index])
                    # Auto-advance to next image
                    time.sleep(0.1)  # Simulate frame rate for images
                    self.current_image_index = (
                        self.current_image_index + 1) % len(self.image_files)
                    return frame

            elif self.current_source_type == 'video':
                if self.cap and self.cap.isOpened():
                    ret, frame = self.cap.read()
                    if not ret:  # End of video, restart
                        self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
                        ret, frame = self.cap.read()
                    return frame if ret else None

        except Exception as e:
            self.log_message(
                f"Error reading frame from {self.current_source_type}: {e}")
        return None

    def _start_display_update(self):
        """Start display update in main thread — keeps running even when minimized"""
        def update_display():
            if not self.running:
                return  # Stop only when system is fully stopped

            try:
                frame = self.frame_queue.get_nowait()
                self.current_frame = frame

                # Only render to screen when window is visible
                window_state = self.root.state()
                if window_state != 'iconic':  # 'iconic' = minimized
                    self._update_display_frame(frame)
                # If minimized, frame is simply discarded after processing —
                # inference and counting already happened in the background thread

            except queue.Empty:
                pass  # No new frame available yet

            # Always reschedule — this keeps counting alive when minimized
            self.root.after(10, update_display)

        update_display()

    def _process_frame(self, frame: np.ndarray) -> Optional[np.ndarray]:
        """Process single frame with optimizations"""
        try:
            # Apply transformations
            frame = self.frame_processor.apply_rotation(
                frame, self.settings.rotate)
            frame = self.frame_processor.resize_frame(
                frame, self.settings.resolution)

            # Run inference (with frame skipping)
            if self.frame_counter % (self.settings.frame_skip + 1) == 0:
                detections = self.inference_engine.infer(
                    frame, self.settings.threshold, self.settings.inference_imgsz)

                # Compute the two independent counting lines
                upper_y = int(frame.shape[0] * self.settings.divider_ratio_upper)
                lower_y = int(frame.shape[0] * self.settings.divider_ratio_lower)

                with self.pause_lock:
                    if not self.counting_paused:
                        newly_counted = self.egg_tracker.update_tracking(
                            detections, upper_y, lower_y)

                        # Handle PLC communication only when not paused
                        self._handle_plc_communication()
                    else:
                        # When paused, we still track detections for display but don't count
                        # This maintains visual tracking while preventing counting
                        newly_counted = 0

                # Draw overlays
                if self.settings.show_overlays:
                    # Use tracks from EggTracker instead of raw detections for stable bounding boxes
                    stable_detections = self.egg_tracker.get_active_tracks()
                    frame = self._draw_detections(frame, stable_detections)
                
                frame = self._draw_interface(frame, upper_y, lower_y)

            # Handle recording - but don't reinitialize when paused
            if hasattr(self, 'record_var') and self.record_var.get():
                if not self.video_recorder.is_recording:
                    # Only start recording once
                    self.video_recorder.start_recording(frame)
                elif self.video_recorder.is_recording:
                    # Just write frame, don't reinitialize
                    self.video_recorder.write_frame(frame)

            # Stream frame (if streaming is active) - ALWAYS stream even when paused
            if hasattr(self, 'mjpeg_streamer') and self.mjpeg_streamer.is_running:
                self.mjpeg_streamer.update_frame(frame)

            return frame

        except Exception as e:
            self.log_message(f"Frame processing error: {e}")
            return None

    def _handle_plc_communication(self):
        """Handle PLC communication"""
        current_time = time.time()

        if (self.egg_tracker.total_eggs_counted != self.last_plc_count and
                current_time - self.last_plc_update_time >= self.plc_update_interval):

            if not self.plc_connection_attempted and self.settings.plc_ip:
                self.plc_connection_attempted = True
                self.connect_plc()

            if self.plc and hasattr(self.plc, 'connected') and self.plc.connected:
                success = self.plc.write_register(
                    self.settings.plc_address, self.egg_tracker.total_eggs_counted)
                if success:
                    self.last_plc_count = self.egg_tracker.total_eggs_counted
                    self.last_plc_update_time = current_time

    def _draw_detections(self, frame: np.ndarray, detections: List[Detection]) -> np.ndarray:
        """Draw detections with robust coloring and unique labeling"""
        # Pre-allocate colors
        regular_color = (150, 150, 150) # Neutral gray for boxes
        counted_text_color = (0, 255, 0) # Vibrant green for count numbers

        for detection in detections:
            xmin, ymin, xmax, ymax = detection.bbox

            # Use different color scheme when paused
            with self.pause_lock:
                if self.counting_paused:
                    box_color = (100, 100, 100) # Faded gray when paused
                    text_color = (100, 200, 100) # Faded green
                else:
                    box_color = regular_color
                    text_color = counted_text_color

            # Draw bounding box
            cv2.rectangle(frame, (xmin, ymin), (xmax, ymax), box_color, 1)
            cv2.circle(frame, detection.center, 3, box_color, -1)

            # Draw the unique Count Number if it exists
            if detection.counted_number is not None:
                self._draw_label(frame, xmin, ymin, xmax, ymax, str(detection.counted_number), text_color)

        return frame



    def _draw_label(self, frame: np.ndarray, xmin: int, ymin: int, xmax: int, ymax: int,
                    count_text: str, color: Tuple[int, int, int]):
        """Draw Count Number above the bounding box with high-contrast shadow"""
        font_scale = 0.65
        thickness = 2
        
        # Position: slightly above the top-left corner
        text_x = xmin
        text_y = max(25, ymin - 12)
        
        # Draw text with 2-layer shadow for maximum industrial visibility
        cv2.putText(frame, count_text, (text_x + 1, text_y + 1),
                    cv2.FONT_HERSHEY_SIMPLEX, font_scale, (0, 0, 0), thickness + 1, cv2.LINE_AA)
        cv2.putText(frame, count_text, (text_x, text_y),
                    cv2.FONT_HERSHEY_SIMPLEX, font_scale, color, thickness, cv2.LINE_AA)


    def _draw_interface(self, frame: np.ndarray, upper_y: int, lower_y: int) -> np.ndarray:
        """Draw interface elements on frame"""
        # Draw two independent counting lines
        with self.pause_lock:
            if self.counting_paused:
                line_color = (255, 0, 0)  # Red when paused
            else:
                line_color = (0, 255, 255)  # Yellow when active

        # Upper border line
        cv2.line(frame, (0, upper_y), (frame.shape[1], upper_y), line_color, 2)
        # Lower border line
        cv2.line(frame, (0, lower_y), (frame.shape[1], lower_y), line_color, 2)

        # Draw statistics
        upper_percentage = int(self.settings.divider_ratio_upper * 100)
        lower_percentage = int(self.settings.divider_ratio_lower * 100)

        # Device info
        device = self.inference_engine.device.upper()
        cv2.putText(frame, f"Device: {device}", (frame.shape[1] - 200, 40),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.4, (100, 255, 100), 1, cv2.LINE_AA)

        # PLC status
        plc_status = f"PLC: {'Connected' if self.plc and hasattr(self.plc, 'connected') and self.plc.connected else 'Disconnected'}"
        plc_color = (0, 255, 0) if (self.plc and hasattr(
            self.plc, 'connected') and self.plc.connected) else (0, 0, 255)
        cv2.putText(frame, plc_status, (frame.shape[1] - 200, 60),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.4, plc_color, 1, cv2.LINE_AA)

        # Count information
        with self.pause_lock:
            count_color = (255, 0, 0) if self.counting_paused else (
                0, 255, 255)

        cv2.putText(frame, f'Total Counted: {self.egg_tracker.total_eggs_counted}', (10, 50),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, count_color, 1, cv2.LINE_AA)
        cv2.putText(frame, f'Current Eggs: {len(self.egg_tracker.egg_tracking)}', (10, 80),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, count_color, 1, cv2.LINE_AA)
        
        cv2.putText(frame, f'FPS: {self.avg_frame_rate:.1f}', (10, 110),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, count_color, 1, cv2.LINE_AA)

        # Role indicator on video
        role_color = (0, 255, 0) if self.is_admin else (255, 165, 0)
        cv2.putText(frame, f'User: {self.role.capitalize()}', (frame.shape[1] - 200, 20),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, role_color, 1, cv2.LINE_AA)

        # Draw pause status overlay
        with self.pause_lock:
            if self.counting_paused:
                # Add semi-transparent overlay
                overlay = frame.copy()
                cv2.rectangle(overlay, (0, 0),
                              (frame.shape[1], frame.shape[0]), (0, 0, 0), -1)
                frame = cv2.addWeighted(overlay, 0.2, frame, 0.8, 0)

                # Draw large pause indicator
                pause_text = "COUNTING PAUSED"
                text_size = cv2.getTextSize(
                    pause_text, cv2.FONT_HERSHEY_SIMPLEX, 2, 3)[0]
                text_x = (frame.shape[1] - text_size[0]) // 2
                text_y = (frame.shape[0] + text_size[1]) // 2

                # Main pause text
                cv2.putText(frame, pause_text, (text_x, text_y),
                            cv2.FONT_HERSHEY_SIMPLEX, 2, (0, 0, 255), 3, cv2.LINE_AA)

                # Subtitle
                subtitle = "Press SPACEBAR to resume "
                subtitle_size = cv2.getTextSize(
                    subtitle, cv2.FONT_HERSHEY_SIMPLEX, 0.7, 2)[0]
                subtitle_x = (frame.shape[1] - subtitle_size[0]) // 2
                cv2.putText(frame, subtitle, (subtitle_x, text_y + 50),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2, cv2.LINE_AA)

        # Divider info
        # cv2.putText(frame, f'Divider: {upper_percentage}% Upper / {lower_percentage}% Lower',
        #             (10, frame.shape[0] - 20), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255, 255, 0), 1, cv2.LINE_AA)

        # System title
        cv2.putText(frame, "DAC Egg Counting System", (frame.shape[1] // 2 - 150, 25),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1, cv2.LINE_AA)

        return frame

    def _update_display_frame(self, frame: np.ndarray):
        """Update display efficiently"""
        try:
            # Convert to RGB
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(frame_rgb)

            # Resize only if canvas size changed
            canvas_width = self.video_canvas.winfo_width()
            canvas_height = self.video_canvas.winfo_height()

            if canvas_width > 1 and canvas_height > 1:
                if (img.width, img.height) != (canvas_width, canvas_height):
                    img = img.resize(
                        (canvas_width, canvas_height), Image.Resampling.LANCZOS)

            # Update photo
            self.photo = ImageTk.PhotoImage(image=img)
            self.video_canvas.delete("all")
            self.video_canvas.create_image(
                0, 0, image=self.photo, anchor=tk.NW)

            # Update performance indicators
            self.fps_var.set(f"{self.avg_frame_rate:.1f}")
            self.total_eggs_var.set(str(self.egg_tracker.total_eggs_counted))
            self.current_eggs_var.set(str(len(self.egg_tracker.egg_tracking)))
            self.tracked_var.set(str(len(self.egg_tracker.egg_tracking)))
            
            # Update detailed classification stats
            self.update_counting_status_display()

        except Exception as e:
            self.log_message(f"Display update error: {e}")

    def handle_keypress(self, event):
        """Handle keyboard shortcuts"""
        if event.char == 'q' or event.char == 'Q':
            self.stop_system()
        elif event.char == 'r' or event.char == 'R':
            self.reset_count()
        elif event.char == 'a' or event.char == 'A':
            self.toggle_adjustment()
        elif event.char == 'p' or event.char == 'P':
            self.capture_image()
        elif event.char == 'g' or event.char == 'G':
            self.toggle_device()
        elif event.char == 'c' or event.char == 'C':
            self.connect_plc()
        elif event.char == 'h' or event.char == 'H' or event.char == 'h' or event.char == 'H':
            # H key toggles control visibility
            self.toggle_controls_visibility()
        elif event.char == 's' or event.char == 'S':
            self.start_streaming()
        elif event.keysym == 'space':
            # Spacebar toggles counting pause - use thread-safe method
            self.toggle_pause_counting()
        elif event.keysym == 'Escape':
            if self.full_screen_mode:
                self.toggle_full_screen()
        elif event.keysym == 'F1':
            self.show_help()

    def show_help(self):
        """Show help information"""
        help_text = f"""
        DAC Egg Counting System - Help
        
        User Role: {self.role.capitalize()}
        Permissions: {'Full Administrative Access' if self.is_admin else 'Operator Access with Full Control Tab'}
        
        Keyboard Shortcuts:
        • SPACEBAR: Pause/Resume counting
        • H: Hide/Show controls
        • F11: Toggle full screen
        • R: Reset count
        • A: Toggle divider adjustment
        • P: Capture image
        • G: Toggle GPU/CPU
        • C: Connect PLC
        • S: Start streaming
        • Q: Stop system
        • ESC: Exit full screen
        
        Control Tab Features:
        • Start/Stop System
        • Load Model
        • Reset Count
        • Adjust Divider
        • Capture Image
        • Pause Counting
        • Toggle GPU/CPU
        • Save Settings
        • Full Screen Toggle
        • Excel/CSV Export
        • Daily Reports
        • Wireless Streaming
        • Quick Settings (Operators)
        """
        messagebox.showinfo("Help", help_text)

    def _pick_divider_line(self, event_y: int) -> str:
        """Return which divider line ('upper' or 'lower') is closer to the click position"""
        canvas_height = self.video_canvas.winfo_height()
        if canvas_height <= 0:
            return 'upper'
        upper_y = int(canvas_height * self.settings.divider_ratio_upper)
        lower_y = int(canvas_height * self.settings.divider_ratio_lower)
        return 'upper' if abs(event_y - upper_y) <= abs(event_y - lower_y) else 'lower'

    def handle_canvas_click(self, event):
        """Handle canvas click for divider adjustment"""
        if self.adjustment_mode:
            canvas_height = self.video_canvas.winfo_height()
            if canvas_height > 0:
                self._dragging_line = self._pick_divider_line(event.y)
                self._move_divider_line(self._dragging_line, event.y, canvas_height)

    def handle_canvas_drag(self, event):
        """Handle canvas drag for divider adjustment"""
        if self.adjustment_mode:
            canvas_height = self.video_canvas.winfo_height()
            if canvas_height > 0:
                line = getattr(self, '_dragging_line', self._pick_divider_line(event.y))
                self._move_divider_line(line, event.y, canvas_height)

    def _move_divider_line(self, line: str, event_y: int, canvas_height: int):
        """Move a specific divider line, keeping upper < lower"""
        new_ratio = max(0.05, min(0.95, event_y / canvas_height))
        if line == 'upper':
            # Upper line must stay above lower line
            new_ratio = min(new_ratio, self.settings.divider_ratio_lower - 0.01)
            self.settings.divider_ratio_upper = new_ratio
            if hasattr(self, 'divider_upper_var'):
                self.divider_upper_var.set(new_ratio)
        else:
            # Lower line must stay below upper line
            new_ratio = max(new_ratio, self.settings.divider_ratio_upper + 0.01)
            self.settings.divider_ratio_lower = new_ratio
            if hasattr(self, 'divider_lower_var'):
                self.divider_lower_var.set(new_ratio)
        self.save_settings()

    def handle_canvas_release(self, event):
        """Handle canvas release"""
        pass

    def toggle_pause_counting(self):
        """Toggle counting pause state - thread-safe version"""
        if not self.running:
            return

        with self.pause_lock:
            self.counting_paused = not self.counting_paused

            if self.counting_paused:
                # Start pause timer
                self.pause_start_time = time.time()
                self.log_message(
                    "⚠️ COUNTING PAUSED - Live feed continues for maintenance")
                self.log_message(
                    "   Eggs detected will not be counted while paused")
                self.log_message("   Press SPACEBAR again to resume counting")

                # Update button text
                self.pause_button.config(text="Resume Counting (Space)")
            else:
                # Calculate pause duration
                if self.pause_start_time:
                    pause_duration = time.time() - self.pause_start_time
                    self.total_pause_time += pause_duration
                    pause_minutes = int(pause_duration // 60)
                    pause_seconds = int(pause_duration % 60)
                    self.log_message(
                        f"✅ Counting RESUMED after {pause_minutes}m {pause_seconds}s pause")
                    self.log_message(
                        f"   Total system pause time: {int(self.total_pause_time // 60)}m {int(self.total_pause_time % 60)}s")
                    self.pause_start_time = None

                # Update button text
                self.pause_button.config(text="Pause Counting")

        # Update status displays
        self.update_pause_status_display()
        self.update_counting_status_display()

    def update_pause_status_display(self):
        """Update the pause status display"""
        with self.pause_lock:
            if self.counting_paused:
                self.pause_status_var.set("Counting: PAUSED")
            else:
                self.pause_status_var.set("Counting: Active")

    def update_counting_status_display(self):
        """Update the counting status in statistics panel"""
        with self.pause_lock:
            if self.counting_paused:
                self.counting_status_var.set("PAUSED")
                self.counting_status_label.config(foreground="red")
            else:
                self.counting_status_var.set("Active")
                self.counting_status_label.config(foreground="green")

            # Update specific class counts
            # Note: We use .get() to handle cases where keys might be capitalized differently or missing
            # Assuming standard naming, but using case-insensitive check safely would be better if unsure.
            # For now, we trust the model output or what we inject.
            
            # Helper to find count case-insensitively and ignoring whitespace/punctuation
            def normalize_name(name):
                return name.lower().replace('_', ' ').replace('-', ' ').strip()
                
            def get_count(target):
                target = normalize_name(target)
                for k, v in self.egg_tracker.class_counts.items():
                    if normalize_name(k) == target:
                        return v
                return 0
                
            self.white_egg_var.set(str(get_count("white egg")))
            self.dirty_egg_var.set(str(get_count("dirty egg")))
            self.broken_egg_var.set(str(get_count("broken egg")))

    def reset_count(self):
        """Reset egg count"""
        self.egg_tracker.reset()
        self.last_plc_count = 0

        self.total_eggs_var.set("0")
        self.current_eggs_var.set("0")
        self.tracked_var.set("0")

        self.log_message("Counting reset!")

    def toggle_adjustment(self):
        """Toggle divider adjustment mode"""
        self.adjustment_mode = not self.adjustment_mode

        if self.adjustment_mode:
            self.log_message(
                "Divider adjustment mode activated - Click and drag on video to adjust")
            self.video_canvas.config(cursor="crosshair")
        else:
            self.log_message("Divider adjustment mode deactivated")
            self.video_canvas.config(cursor="")

    def capture_image(self):
        """Capture current frame as image"""
        if self.current_frame is not None:
            filename = f"capture_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            cv2.imwrite(filename, self.current_frame)
            self.log_message(f"Image captured: {filename}")

    def open_excel_file(self):
        """Open the Excel file in default application"""
        try:
            if os.path.exists(self.excel_exporter.filename):
                os.startfile(self.excel_exporter.filename)  # Windows
                self.log_message("Opening Excel file...")
            else:
                self.log_message(
                    "Excel file not created yet - will be created after first minute")
        except Exception as e:
            self.log_message(f"Error opening Excel: {e}")

    def create_daily_report(self):
        """Manual daily report generation"""
        report_file = self.csv_exporter.create_daily_report()
        if report_file:
            self.log_message(f"Daily report created: {report_file}")
            messagebox.showinfo(
                "Report Created", f"Daily report saved as: {report_file}")
        else:
            messagebox.showwarning(
                "Report Failed", "Could not create daily report")
        return report_file

    def show_export_status(self):
        """Show export file status"""
        excel_exists = os.path.exists(self.excel_exporter.filename)
        csv_exists = os.path.exists(self.csv_exporter.filename)

        with self.pause_lock:
            status_msg = f"User Role: {self.role.capitalize()}\n"
            status_msg += f"Excel file: {'Exists' if excel_exists else 'Not created'}\n"
            status_msg += f"CSV file: {'Exists' if csv_exists else 'Not created'}\n"
            status_msg += f"Total eggs: {self.egg_tracker.total_eggs_counted}\n"
            status_msg += f"Counting status: {'PAUSED' if self.counting_paused else 'ACTIVE'}\n"
            status_msg += f"Streaming: {'Active' if self.mjpeg_streamer.is_running else 'Stopped'}"

        messagebox.showinfo("Export Status", status_msg)

    def show_stream_info(self):
        """Show streaming connection information with robust IP detection"""
        local_ip = get_robust_local_ip()

        stream_info = f"🌐 WIRELESS STREAMING INFORMATION\n\n"
        stream_info += f"Web Stream URL:\n"
        stream_info += f"http://{local_ip}:8080\n\n"
        stream_info += f"Direct Stream URL:\n"
        stream_info += f"http://{local_ip}:8080/stream.mjpg\n\n"
        stream_info += "📱 How to Connect:\n"
        stream_info += "• Phones/Tablets: Open web browser\n"
        stream_info += "• Computers: Use any web browser\n"
        stream_info += "• No software installation required!\n\n"
        stream_info += "✅ Works on all devices automatically"

        messagebox.showinfo("Streaming Information", stream_info)

    def stop_system(self):
        """Stop system efficiently"""
        if not self.running:
            return
            
        self.running = False
        self._update_button_states()

        # Stop exporters and streamers
        self.csv_exporter.stop()
        self.excel_exporter.stop()
        self.stop_streaming()

        # Clean up resources
        if self.cap:
            self.cap.release()
            self.cap = None

        if self.delta_cam:
            self.delta_cam.disconnect()
            self.delta_cam = None

        self.video_recorder.stop_recording()

        if self.plc and hasattr(self.plc, 'connected') and self.plc.connected:
            self.plc.write_register(
                self.settings.plc_address, self.egg_tracker.total_eggs_counted)
            self.plc.disconnect()

        self.system_status_var.set(
            f"System Stopped - {self.role.capitalize()} Mode")
        self.log_message("Egg counting system stopped")


class DeltaCamera:
    def __init__(self, cti_path=r"C:\Program Files\Delta Industrial Automation\DIAVision\DMV-SDK\Runtime\x64\dmvc-producer.cti"):
        self.cti_path = cti_path
        self.h = None
        self.ia = None
        self.connected = False
        self.gui = None

    def connect(self):
        """Connect to Delta camera"""
        try:
            from harvesters.core import Harvester

            self.h = Harvester()

            try:
                self.h.add_file(self.cti_path)
                self.log_message("✓ Delta GenTL producer loaded")
            except Exception as e:
                self.log_message(f"✗ Failed to load CTI: {e}")
                return False

            try:
                self.h.update()
                self.log_message(
                    f"Found {len(self.h.device_info_list)} device(s)")

                if len(self.h.device_info_list) == 0:
                    self.log_message("✗ No Delta cameras found")
                    return False

                # Create image acquirer
                self.ia = self.h.create_image_acquirer(0)

                self.log_message("Connected to: DMV-CC400GC290")

                # Display camera information
                try:
                    model = self.ia.remote_device.node_map.DeviceModelName.value
                    serial = self.ia.remote_device.node_map.DeviceSerialNumber.value
                    self.log_message(f"Camera Model: {model}")
                    self.log_message(f"Serial Number: {serial}")
                except Exception as e:
                    self.log_message(
                        f"Camera info: DMV-CC400GC290 (C400GC0D23190042)")

                self.connected = True
                return True

            except Exception as e:
                self.log_message(f"Error connecting to camera: {e}")
                return False

        except ImportError:
            self.log_message(
                "✗ Harvesters library not available for Delta camera")
            return False

    def start_acquisition(self):
        """Start camera acquisition"""
        if self.connected and self.ia:
            self.ia.start()
            self.log_message("Delta camera acquisition started")

    def stop_acquisition(self):
        """Stop camera acquisition"""
        if self.connected and self.ia:
            try:
                self.ia.stop()
            except:
                pass

    def read_frame(self):
        """Read a frame from Delta camera"""
        if not self.connected or not self.ia:
            return False, None

        try:
            with self.ia.fetch(timeout=5000) as buffer:
                # Get image data
                component = buffer.payload.components[0]
                img_data = component.data
                height = component.height
                width = component.width

                # Reshape and convert image
                img = img_data.reshape(height, width)

                # Try different color conversions
                try:
                    display_img = cv2.cvtColor(img, cv2.COLOR_BAYER_BG2BGR)
                except:
                    try:
                        display_img = cv2.cvtColor(img, cv2.COLOR_BAYER_RG2BGR)
                    except:
                        display_img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)

                return True, display_img

        except Exception as e:
            self.log_message(f"Delta camera acquisition error: {e}")
            return False, None

    def disconnect(self):
        """Disconnect from Delta camera"""
        self.stop_acquisition()
        if self.ia:
            try:
                self.ia.destroy()
            except:
                pass
        if self.h:
            try:
                self.h.reset()
            except:
                pass
        self.connected = False
        self.log_message("Delta camera disconnected")

    def log_message(self, message):
        """Log message if parent GUI exists"""
        if self.gui:
            self.gui.log_message(message)
        else:
            print(message)


class ModbusTCPPLC:
    def __init__(self, ip='169.254.163.50', port=502, unit_id=1, timeout=3):
        self.ip = ip
        self.port = port
        self.unit_id = unit_id
        self.timeout = timeout
        self.client = None
        self.connected = False
        self.gui = None

    def connect(self):
        """Connect to PLC via MODBUS TCP"""
        try:
            self.client = ModbusClient(
                host=self.ip,
                port=self.port,
                unit_id=self.unit_id,
                timeout=self.timeout,
                auto_open=True,
                auto_close=False
            )

            # Test connection by reading a register
            if self.client.open():
                self.connected = True
                self.log_message(
                    f"✅ Successfully connected to PLC at {self.ip}:{self.port}")
                self.log_message(f"📡 MODBUS TCP - Unit ID: {self.unit_id}")
                return True
            else:
                self.log_message(
                    f"❌ Failed to connect to PLC at {self.ip}:{self.port}")
                return False

        except Exception as e:
            self.log_message(
                f"❌ Connection error to PLC {self.ip}:{self.port}: {e}")
            return False

    def disconnect(self):
        """Disconnect from PLC"""
        if self.client:
            self.client.close()
            self.connected = False
            self.log_message("Disconnected from PLC")

    def write_register(self, address, value):
        """Write value to holding register"""
        try:
            # Ensure value is within 0-65535 (16-bit)
            value = max(0, min(65535, value))

            # Write to holding register
            success = self.client.write_single_register(address, value)

            if success:
                self.log_message(
                    f"✅ MODBUS TCP: Successfully wrote {value} to register D{address}")
                return True
            else:
                self.log_message(
                    f"❌ MODBUS TCP: Failed to write to register D{address}")
                return False

        except Exception as e:
            self.log_message(f"❌ MODBUS TCP communication error: {e}")
            return False

    def read_register(self, address):
        """Read value from holding register"""
        try:
            value = self.client.read_holding_registers(address, 1)
            if value:
                return value[0]
            else:
                self.log_message(
                    f"❌ MODBUS TCP: Failed to read register D{address}")
                return None
        except Exception as e:
            self.log_message(f"❌ MODBUS TCP read error: {e}")
            return None

    def test_communication(self):
        """Test basic communication with PLC"""
        if not self.connected:
            return False

        try:
            # Try to read a register to test communication
            test_value = self.read_register(100)
            if test_value is not None:
                self.log_message(
                    f"🔧 PLC communication test successful - Register D100 value: {test_value}")
                return True
            else:
                self.log_message(
                    "🔧 PLC communication test: Could not read register (may be normal)")
                return True

        except Exception as e:
            self.log_message(f"❌ PLC test communication error: {e}")
            return False

    def log_message(self, message):
        """Log message if parent GUI exists"""
        if self.gui:
            self.gui.log_message(message)
        else:
            print(message)


def main():
    """Main function with error handling"""
    initialize_frozen_environment()
    app_instance = None
    try:
        root = tk.Tk()
        app = LoginApp(root)

        # Keep reference to check later
        def check_app():
            nonlocal app_instance
            if hasattr(app, 'main_app'):
                app_instance = app.main_app
            root.after(1000, check_app)

        root.after(1000, check_app)
        root.mainloop()

    except Exception as e:
        print(f"CRITICAL ERROR: {e}")
        import traceback
        traceback.print_exc()

        # Emergency save attempt
        if app_instance:
            print("Attempting emergency save...")
            try:
                if hasattr(app_instance, 'stop_system'):
                    app_instance.stop_system()
                print("Emergency save completed.")
            except Exception as save_err:
                print(f"Failed to perform emergency save: {save_err}")

        with open("crash_log.txt", "a") as f:
            f.write(f"\n[{datetime.datetime.now()}] Crash Report:\n")
            traceback.print_exc(file=f)

        messagebox.showerror(
            "Critical Error", f"Application crashed via main loop.\nCheck crash_log.txt\nError: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
